# backend/scripts/build_timesheet_template.py
"""Sanitize the June 2026 attendance workbook into the reusable template.

June is the source because it still has the company logo — the July attendance
file on the share lost its image to an outside tool. One-off; re-run only if the
client changes the paper.

    python backend/scripts/build_timesheet_template.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.worksheet.datavalidation import DataValidationList

SOURCE = Path(
    r"E:\Al Watbha Shares\المالية\احصائية 2026\6-Jun\كشف حضور شهر يونيو.xlsx"
)
DEST = Path(__file__).resolve().parents[1] / "templates" / "GSSG-HR_Monthly_Time_Sheet.xlsx"

FIRST_DATA_ROW = 6
# June row 6 is the only row carrying the header-boundary rule -- a medium top border
# in AK..AP -- and row 5's bottom border is None there, so that rule is the only thing
# drawing the line under the header in those six columns. Copied into every output row
# it would draw ~280 times; dropped, the boundary is lost. Hence two specimens. Row 7 is
# the representative body row: it differs from row 6 in exactly AK..AP and nowhere else,
# and 100 of June's 282 data rows match its full 42-column style (row 6 matches only itself).
SUBSEQUENT_DATA_ROW = 7
LAST_DATA_ROW = 287  # June's last employee row; its footer starts at 288
FOOTER_ROWS = 18  # legend .. Total Days
LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def main() -> None:
    workbook = load_workbook(SOURCE)
    sheet = workbook.worksheets[0]
    assert sheet.title == "Sheet1", sheet.title  # already named this in the source

    parts = workbook.create_sheet("_parts")
    parts.sheet_state = "hidden"

    # _parts row 1: the first-data-row specimen (header boundary), values stripped.
    for column in range(1, 43):
        parts.cell(1, column)._style = sheet.cell(FIRST_DATA_ROW, column)._style
    parts.row_dimensions[1].height = sheet.row_dimensions[FIRST_DATA_ROW].height

    # _parts row 2: the subsequent-row specimen, for every output row after the first.
    for column in range(1, 43):
        parts.cell(2, column)._style = sheet.cell(SUBSEQUENT_DATA_ROW, column)._style
    parts.row_dimensions[2].height = sheet.row_dimensions[SUBSEQUENT_DATA_ROW].height

    # _parts rows 3..20: the footer block, styles and static text intact.
    footer_start = LAST_DATA_ROW + 1
    for offset in range(FOOTER_ROWS):
        for column in range(1, 43):
            source_cell = sheet.cell(footer_start + offset, column)
            target = parts.cell(3 + offset, column)
            target._style = source_cell._style
            if isinstance(source_cell.value, str) and not source_cell.value.startswith("="):
                target.value = source_cell.value
        height = sheet.row_dimensions[footer_start + offset].height
        if height:
            parts.row_dimensions[3 + offset].height = height

    # Strip Sheet1 back to the header. Unmerge FIRST — a MergedCell's .value is
    # read-only — then pop the cells so no style index survives.
    for merged in [str(r) for r in sheet.merged_cells.ranges]:
        if int(merged.split(":")[0].lstrip(LETTERS)) >= FIRST_DATA_ROW:
            sheet.unmerge_cells(merged)
    for row in range(FIRST_DATA_ROW, sheet.max_row + 1):
        sheet.row_dimensions.pop(row, None)
        for column in range(1, 43):
            sheet._cells.pop((row, column), None)

    # Rebuilt from scratch by the renderer over the real extent. June carries FOUR
    # conditionalFormatting blocks of nine `cellIs equal` rules each (36 rules, 36
    # dxfs) over a heavily fragmented sqref -- "D299 F6:AI28 F29:AJ29 ... F225:AI287"
    # plus the strays AJ88, AJ248, AJ257 -- and THREE list validations whose
    # formula1 is "$D$295:$D$304", a reference to the footer code rows this strip
    # deletes. Copying any of it forward ships dangling references.
    sheet.conditional_formatting = ConditionalFormattingList()
    sheet.data_validations = DataValidationList()

    # June's autofilter is "D5:E290" -- a range over rows the strip just deleted, so
    # every rendered month would carry dropdowns frozen at June's extent. Its four
    # `.wvu.FilterData` names are sheet-local on Sheet1 (the workbook-level
    # collection is empty) and are orphans: openpyxl drops the customSheetViews they
    # point at on save. Same class of June artifact as freeze_panes="A77".
    sheet.auto_filter.ref = None
    sheet.defined_names.clear()

    # The 19th footer row: the manual red block X, inserted after OFF (parts row
    # 19) and before Total Days, which moves from row 20 to 21. Copy downward
    # first or the Total Days values are overwritten.
    for column in range(1, 43):
        parts.cell(21, column)._style = parts.cell(20, column)._style
        parts.cell(21, column).value = parts.cell(20, column).value
        parts.cell(20, column)._style = parts.cell(19, column)._style
        parts.cell(20, column).value = None
    parts.row_dimensions[21].height = parts.row_dimensions[20].height
    parts.row_dimensions[20].height = parts.row_dimensions[19].height  # OFF, not Total Days
    parts["C20"].value = "Not billed"
    parts["D20"].value = "X"
    parts["A3"].value = f"{parts['A3'].value}, X- Not billed"

    sheet["D4"].value = "For the Month of :"
    sheet.freeze_panes = "F6"  # June's source is "A77", a mid-roster artifact

    DEST.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(DEST)
    print(f"[template] wrote {DEST} ({DEST.stat().st_size} bytes)")

    check = load_workbook(DEST)
    assert len(check["Sheet1"]._images) == 1, "logo lost"
    assert check["Sheet1"].max_row == 5, f"stray rows: max_row={check['Sheet1'].max_row}"
    assert check["_parts"]["A21"].value == "Total Days", "footer is not 19 rows"
    assert check["_parts"]["D20"].value == "X", "red block row missing"
    assert check["Sheet1"].auto_filter.ref is None, "June's autofilter survived"
    assert not list(check["Sheet1"].defined_names), "orphaned filter names survived"
    specimens = check["_parts"]
    differing = {c for c in range(1, 43) if specimens.cell(1, c)._style != specimens.cell(2, c)._style}
    assert differing == set(range(37, 43)), f"specimens differ outside AK..AP: {sorted(differing)}"
    print("[template] logo, strip, autofilter, two specimens and 19-row footer verified")


if __name__ == "__main__":
    main()
