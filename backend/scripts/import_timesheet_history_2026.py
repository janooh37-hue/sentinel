"""Reconcile the 2026 hand-kept attendance workbooks into the database.

The monthly attendance grid for JD 908 was maintained by hand in Excel all
through 2026. Most of it is already in the DB — annual/sick/national-service
leave, join dates, departures — but some leave was never recorded, some rows
were saved with a lost type, and two employee records are wrong. This script
closes that gap so the app can reproduce June and July 2026 from the DB before
the time-sheet feature is built on that assumption.

Everything is derived from the workbooks at run time, never from a frozen list:
the eight 2026 attendance sheets are loaded, a continuous day → code series is
built per employee, and maximal runs are taken per code. Runs are therefore
correctly bounded even when a leave crosses a month edge.

Planned writes (see the spec for the full tables):

* create the annual / sick / national-service leave records the sheets show and
  the DB lacks, for any run touching June or July 2026
* create absence records for every ``AB`` run in those two months
* correct three leave rows the sheets contradict (G3101, G3190, G3209)
* retype the surviving ``Unknown`` rows to annual leave, and delete the two that
  merely duplicate a typed row
* fix ``G4537.end_date`` (recorded a year out) and merge the duplicate employee
  ``5704`` into ``G5704``

DRY-RUN by default — prints the plan and exits without writing. ``--apply``
mutates and ONLY after a fresh backup; ``--verify`` regenerates June and July
from the DB and diffs every cell against the workbooks.

    python backend/scripts/import_timesheet_history_2026.py
    python backend/scripts/import_timesheet_history_2026.py --apply
    python backend/scripts/import_timesheet_history_2026.py --verify

Reference: docs/superpowers/specs/2026-08-19-monthly-timesheet-design.md
"""

from __future__ import annotations

import argparse
import calendar
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session, sessionmaker

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))  # put backend/ on the path

from app.core.leave_lifecycle import english_part
from app.core.timesheet_codes import (
    CODE_ABSENT,
    CODE_ANNUAL,
    CODE_NATIONAL,
    CODE_SICK,
    LeaveSpan,
    in_roster,
    is_void,
    leave_code,
    month_codes,
)
from app.db.session import _sqlite_url_for, attach_sqlite_pragmas

DEFAULT_DB = Path(__file__).resolve().parents[2] / "data" / "gssg.db"
SHARE = Path(r"E:\Al Watbha Shares\المالية\احصائية 2026")

#: The 2026 attendance workbooks, newest-complete file per month. The
#: ``...23...`` siblings hold only days 23-31 and are deliberately skipped.
MONTH_FILES: dict[int, str] = {
    1: r"1-January\كشف حضور شهر يناير .xlsx",
    2: r"2-February\كشف حضور شهر فبراير.xlsx",
    3: r"3-March\كشف حضور شهر مارس.xlsx",
    4: r"4- April\كشف الحضور شهر ابريل .xlsx",
    5: r"5-May\كشف حضور شهر مايو.xlsx",
    6: r"6-Jun\كشف حضور شهر يونيو.xlsx",
    7: r"7-Jul\كشف حضور شهر يوليو.xlsx",
    8: r"8-Aug\كشف حضور شهر اغسطس_backup_20260803_101702.xlsx",
}

#: The drivers roster lives in its own workbook, never in the main sheet.
DRIVERS_FILE = r"7-Jul\كشف حضور شهر يوليو للسائقين.xlsx"

FIRST_DATA_ROW = 6
FIRST_DAY_COL = 6  # column F
ID_COL = 2
DESIGNATION_COL = 5  # column E

#: The months the import must make reproducible.
TARGET = (date(2026, 6, 1), date(2026, 7, 31))

#: Sheet code → the leave type created for it.
RUN_TYPES = {
    CODE_ANNUAL: "Annual Leave",
    CODE_SICK: "Sick Leave",
    CODE_NATIONAL: "National Service",
}

#: Leave rows the workbooks contradict: drop the row, create the replacement.
CORRECTIONS: tuple[tuple[str, str, str, str, str | None, str | None, str | None], ...] = (
    # employee, type, start, end, new_type, new_start, new_end  (None new_type = delete only)
    ("G3101", "Unknown", "2026-06-17", "2026-06-30", None, None, None),
    ("G3101", "Unknown", "2026-06-17", "2026-07-04", "Annual Leave", "2026-06-19", "2026-06-30"),
    ("G3190", "Sick Leave", "2026-06-19", "2026-06-22", "Sick Leave", "2026-06-19", "2026-06-19"),
    ("G3209", "Unknown", "2026-07-01", "2026-07-30", "Annual Leave", "2026-07-01", "2026-07-26"),
)

#: Employee records the workbooks prove wrong.
END_DATE_FIXES: dict[str, str] = {"G4537": "2026-06-17"}

#: Duplicate employee: (drop, keep). Identical name/doj/nationality, no leaves.
DUPLICATE_EMPLOYEE = ("5704", "G5704")


@dataclass(frozen=True, slots=True)
class Run:
    """A maximal run of one code for one employee, read off the workbooks."""

    employee_id: str
    code: str
    start: date
    end: date

    @property
    def days(self) -> int:
        return (self.end - self.start).days + 1


def _session_for(db_path: Path) -> Session:
    eng = create_engine(_sqlite_url_for(str(db_path)), future=True)
    attach_sqlite_pragmas(eng, wal=False)
    return sessionmaker(bind=eng, future=True, expire_on_commit=False)()


def canonical_employee_id(employee_id: str) -> str:
    """Map the known duplicate G-number onto the surviving record.

    Applied at the readers so nothing downstream — leave runs, absence runs,
    designation links — can be keyed to a record this run deletes.
    """

    drop, keep = DUPLICATE_EMPLOYEE
    return keep if employee_id == drop else employee_id


def read_series() -> dict[str, dict[date, str]]:
    """Load every 2026 workbook into ``{employee_id: {date: code}}``."""

    series: dict[str, dict[date, str]] = defaultdict(dict)
    for month, rel in MONTH_FILES.items():
        path = SHARE / rel
        if not path.exists():
            raise SystemExit(f"missing workbook: {path}")
        # read_only makes ws.cell() O(rows) per call; iter_rows streams instead.
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheet = workbook.worksheets[0]
            days = calendar.monthrange(2026, month)[1]
            last_col = FIRST_DAY_COL + days - 1
            for row in sheet.iter_rows(min_row=FIRST_DATA_ROW, max_col=last_col, values_only=True):
                raw_id = row[ID_COL - 1]
                if raw_id is None or not str(raw_id).strip():
                    break
                employee_id = canonical_employee_id(str(raw_id).strip())
                for offset in range(days):
                    value = row[FIRST_DAY_COL - 1 + offset]
                    if value is None:
                        continue
                    code = str(value).strip().upper()
                    if code:
                        series[employee_id][date(2026, month, offset + 1)] = code
        finally:
            workbook.close()
    return dict(series)


def maximal_runs(series: dict[str, dict[date, str]], code: str) -> list[Run]:
    """Every maximal consecutive run of ``code``, across month boundaries."""

    wanted = code.strip().upper()
    runs: list[Run] = []
    for employee_id, by_day in series.items():
        days = sorted(day for day, value in by_day.items() if value == wanted)
        start = previous = None
        for day in days:
            if previous is not None and (day - previous).days == 1:
                previous = day
                continue
            if start is not None and previous is not None:
                runs.append(Run(employee_id, code, start, previous))
            start = previous = day
        if start is not None and previous is not None:
            runs.append(Run(employee_id, code, start, previous))
    return runs


def load_db_state(db: Session) -> tuple[dict[str, dict[str, object]], list[dict[str, object]]]:
    employees = {
        row.id: {
            "name_en": row.name_en,
            "status": row.status,
            "doj": row.doj,
            "end_date": row.end_date,
            "designation_id": row.designation_id,
        }
        for row in db.execute(
            text("SELECT id, name_en, status, doj, end_date, designation_id FROM employees")
        ).all()
    }
    leaves = [
        {
            "id": row.id,
            "employee_id": row.employee_id,
            "leave_type": row.leave_type,
            "start_date": row.start_date,
            "end_date": row.end_date,
            "status": row.status,
        }
        for row in db.execute(
            text(
                "SELECT id, employee_id, leave_type, start_date, end_date, status "
                "FROM leaves WHERE deleted_at IS NULL"
            )
        ).all()
    ]
    return employees, leaves


def _as_date(value: object) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value)[:10])


def covered_days(leaves: list[dict[str, object]], employee_id: str, code: str) -> set[date]:
    """Days already covered by a DB leave that resolves to ``code``."""

    covered: set[date] = set()
    for row in leaves:
        if row["employee_id"] != employee_id or is_void(str(row["status"])):
            continue
        if leave_code(str(row["leave_type"])) != code:
            continue
        start, end = _as_date(row["start_date"]), _as_date(row["end_date"])
        if start is None or end is None:
            continue
        for offset in range((end - start).days + 1):
            covered.add(start + timedelta(days=offset))
    return covered


def read_designations() -> dict[str, str]:
    """``{employee_id: designation_en}``, newest workbook wins.

    Walked newest-first so a promotion recorded in a later month beats an older
    sheet, and the drivers workbook is read last because its two rows appear in
    no other file.
    """

    found: dict[str, str] = {}
    sources = [MONTH_FILES[month] for month in sorted(MONTH_FILES, reverse=True)]
    sources.append(DRIVERS_FILE)
    for rel in sources:
        path = SHARE / rel
        if not path.exists():
            raise SystemExit(f"missing workbook: {path}")
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheet = workbook.worksheets[0]
            for row in sheet.iter_rows(
                min_row=FIRST_DATA_ROW, max_col=DESIGNATION_COL, values_only=True
            ):
                raw_id = row[ID_COL - 1]
                if raw_id is None or not str(raw_id).strip():
                    break
                employee_id = canonical_employee_id(str(raw_id).strip())
                raw_name = row[DESIGNATION_COL - 1]
                if employee_id in found or raw_name is None:
                    continue
                name = str(raw_name).strip()
                if name:
                    found[employee_id] = name
        finally:
            workbook.close()
    return found


def employee_tables(db: Session) -> list[str]:
    """Every table carrying an ``employee_id`` column, discovered not hard-coded.

    Discovery rather than a literal list so a table added later cannot leave an
    orphan row behind when the duplicate employee is merged away.
    """

    names = [
        str(row[0])
        for row in db.execute(
            text("SELECT name FROM sqlite_master WHERE type = 'table' ORDER BY name")
        ).all()
    ]
    # Frozen deliverables are excluded on purpose: a snapshot row reproduces a
    # sheet the client already holds, G-number and all, and must not be rewritten
    # by a later merge (see TimesheetSnapshotRow in app/db/models.py).
    frozen = {"timesheet_snapshot_rows"}
    carriers = []
    for name in names:
        if name in frozen:
            continue
        columns = {str(row[1]) for row in db.execute(text(f'PRAGMA table_info("{name}")')).all()}
        if "employee_id" in columns:
            carriers.append(name)
    return carriers


def employee_reference_counts(db: Session, employee_id: str) -> dict[str, int]:
    """Rows pointing at ``employee_id``, per table, skipping empties."""

    counts: dict[str, int] = {}
    for table in employee_tables(db):
        found = db.execute(
            text(f'SELECT COUNT(*) FROM "{table}" WHERE employee_id = :i'), {"i": employee_id}
        ).scalar_one()
        if found:
            counts[table] = int(found)
    return counts


@dataclass(frozen=True, slots=True)
class Plan:
    new_leaves: list[Run]
    new_absences: list[Run]
    corrections: list[tuple[int, str, str, str, str, str | None, str | None, str | None]]
    retypes: list[tuple[int, str, str, str]]
    deletions: list[tuple[int, str, str, str]]
    end_date_fixes: list[tuple[str, str, str | None]]
    duplicate: tuple[str, str, str | None] | None
    #: table → rows on the dropped employee that must be re-pointed first.
    duplicate_refs: dict[str, int]
    #: (employee_id, designation_id, name_en) rows to set.
    designations: list[tuple[str, int, str]]
    #: employee_id → sheet designation with no catalog match.
    unmatched_designations: dict[str, str]


def build_plan(series: dict[str, dict[date, str]], db: Session) -> Plan:
    employees, leaves = load_db_state(db)
    corrected_keys = {(c[0], c[1], c[2], c[3]) for c in CORRECTIONS}

    new_leaves: list[Run] = []
    for code in RUN_TYPES:
        for run in maximal_runs(series, code):
            if run.employee_id not in employees:
                continue
            if run.end < TARGET[0] or run.start > TARGET[1]:
                continue
            have = covered_days(leaves, run.employee_id, code)
            span = {run.start + timedelta(days=n) for n in range(run.days)}
            if span - have:
                new_leaves.append(run)
    new_leaves.sort(key=lambda r: (RUN_TYPES[r.code], -r.days, r.employee_id))

    recorded_absences = {
        (str(row.employee_id), _as_date(row.date))
        for row in db.execute(text("SELECT employee_id, date FROM absences")).all()
    }
    new_absences = [
        run
        for run in maximal_runs(series, CODE_ABSENT)
        if run.employee_id in employees
        and not (run.end < TARGET[0] or run.start > TARGET[1])
        and any(
            (run.employee_id, run.start + timedelta(days=n)) not in recorded_absences
            for n in range(run.days)
        )
    ]
    new_absences.sort(key=lambda r: (r.start, r.employee_id))

    by_key: dict[tuple[str, str, str, str], int] = {}
    for row in leaves:
        key = (
            str(row["employee_id"]),
            english_part(str(row["leave_type"])),
            str(row["start_date"])[:10],
            str(row["end_date"])[:10],
        )
        by_key.setdefault(key, int(str(row["id"])))

    corrections = [
        (by_key[(c[0], c[1], c[2], c[3])], *c)
        for c in CORRECTIONS
        if (c[0], c[1], c[2], c[3]) in by_key
    ]

    retypes: list[tuple[int, str, str, str]] = []
    deletions: list[tuple[int, str, str, str]] = []
    for row in leaves:
        if english_part(str(row["leave_type"])) != "Unknown":
            continue
        employee_id = str(row["employee_id"])
        start, end = str(row["start_date"])[:10], str(row["end_date"])[:10]
        if (employee_id, "Unknown", start, end) in corrected_keys or is_void(str(row["status"])):
            continue
        typed = [
            r
            for r in leaves
            if r["employee_id"] == employee_id
            and english_part(str(r["leave_type"])) == "Annual Leave"
            and _as_date(r["start_date"]) is not None
            and _as_date(r["start_date"]) <= _as_date(end)  # type: ignore[operator]
            and _as_date(r["end_date"]) >= _as_date(start)  # type: ignore[operator]
        ]
        target = deletions if typed else retypes
        target.append((int(str(row["id"])), employee_id, start, end))

    end_date_fixes = [
        (employee_id, new, str(employees[employee_id]["end_date"] or "") or None)
        for employee_id, new in END_DATE_FIXES.items()
        if employee_id in employees
        and str(employees[employee_id]["end_date"] or "")[:10] != new  # already correct
    ]

    drop, keep = DUPLICATE_EMPLOYEE
    duplicate = (
        (drop, keep, str(employees[drop]["end_date"] or "") or None)
        if drop in employees and keep in employees
        else None
    )
    duplicate_refs = employee_reference_counts(db, drop) if duplicate else {}

    catalog = {
        str(row.name_en).strip().casefold(): int(row.id)
        for row in db.execute(text("SELECT id, name_en FROM timesheet_designations")).all()
    }
    designations: list[tuple[str, int, str]] = []
    unmatched: dict[str, str] = {}
    for employee_id, name in read_designations().items():
        record = employees.get(employee_id)
        if record is None:
            continue
        designation_id = catalog.get(name.casefold())
        if designation_id is None:
            unmatched[employee_id] = name
            continue
        if record.get("designation_id") != designation_id:
            designations.append((employee_id, designation_id, name))
    designations.sort()

    return Plan(
        new_leaves,
        new_absences,
        corrections,
        retypes,
        deletions,
        end_date_fixes,
        duplicate,
        duplicate_refs,
        designations,
        unmatched,
    )


def print_plan(plan: Plan) -> None:
    kinds = Counter(RUN_TYPES[r.code] for r in plan.new_leaves)
    print("[import] leave records to create:")
    for kind, count in sorted(kinds.items()):
        days = sum(r.days for r in plan.new_leaves if RUN_TYPES[r.code] == kind)
        print(f"           {count:3} x {kind:16} ({days} days)")
    for run in plan.new_leaves:
        print(
            f"             {run.employee_id:7} {RUN_TYPES[run.code]:16} {run.start} .. {run.end}  {run.days:3}d"
        )
    outside = [r for r in plan.new_leaves if r.start < TARGET[0] or r.end > TARGET[1]]
    if outside:
        print(
            f"[import] NOTE {len(outside)} runs extend outside {TARGET[0]}..{TARGET[1]}. "
            f"They are written in full (a leave really does cross a month edge) but "
            f"--verify only diffs June and July, so these tails are unverified:"
        )
        for run in outside:
            print(
                f"             {run.employee_id:7} {RUN_TYPES[run.code]:16} {run.start} .. {run.end}"
            )
    print(
        f"[import] absence records to create: {len(plan.new_absences)} runs, "
        f"{sum(r.days for r in plan.new_absences)} days"
    )
    for run in plan.new_absences:
        print(f"             {run.employee_id:7} {run.start} .. {run.end}  {run.days}d")
    print(f"[import] leave rows to correct: {len(plan.corrections)}")
    for (
        leave_id,
        emp,
        old_type,
        old_start,
        old_end,
        new_type,
        new_start,
        new_end,
    ) in plan.corrections:
        after = f"{new_type} {new_start}..{new_end}" if new_type else "delete"
        print(f"             id={leave_id} {emp:7} {old_type} {old_start}..{old_end}  ->  {after}")
    print(f"[import] 'Unknown' rows to retype as Annual Leave: {len(plan.retypes)}")
    for leave_id, emp, start, end in plan.retypes:
        print(f"             id={leave_id} {emp:7} {start}..{end}")
    print(f"[import] 'Unknown' duplicates to delete: {len(plan.deletions)}")
    for leave_id, emp, start, end in plan.deletions:
        print(f"             id={leave_id} {emp:7} {start}..{end}")
    print(f"[import] designation links to set: {len(plan.designations)}")
    by_name = Counter(name for _e, _i, name in plan.designations)
    for name, count in sorted(by_name.items(), key=lambda kv: (-kv[1], kv[0])):
        print(f"             {count:4} x {name}")
    if plan.unmatched_designations:
        print(
            f"[import] WARNING sheet designations with no catalog match: "
            f"{len(plan.unmatched_designations)}"
        )
        for employee_id, name in sorted(plan.unmatched_designations.items()):
            print(f"             {employee_id:7} {name!r}")
    print(f"[import] employee end_date fixes: {len(plan.end_date_fixes)}")
    for employee_id, new, old in plan.end_date_fixes:
        print(f"             {employee_id:7} end_date {old} -> {new}")
    if plan.duplicate:
        drop, keep, end = plan.duplicate
        print(
            f"[import] duplicate employee: merge {drop} -> {keep} "
            f"(end_date {end}, status Resigned), delete {drop}"
        )
        if plan.duplicate_refs:
            total = sum(plan.duplicate_refs.values())
            print(f"             re-point {total} dependent rows from {drop} to {keep}:")
            for table, count in sorted(plan.duplicate_refs.items()):
                print(f"               {count:4} x {table}")
        else:
            print(f"             no dependent rows on {drop}")


def apply_plan(db: Session, plan: Plan) -> None:
    now = datetime.now(UTC).replace(tzinfo=None, microsecond=0).isoformat(sep=" ")
    note = "imported from the 2026 attendance workbooks (timesheet history import)"

    for run in plan.new_leaves:
        db.execute(
            text(
                "INSERT OR IGNORE INTO leaves (employee_id, leave_type, start_date, end_date, "
                "days, status, notes, created_at) "
                "VALUES (:e, :t, :s, :d, :n, 'Approved', :note, :now)"
            ),
            {
                "e": run.employee_id,
                "t": RUN_TYPES[run.code],
                "s": run.start.isoformat(),
                "d": run.end.isoformat(),
                "n": run.days,
                "note": note,
                "now": now,
            },
        )

    for run in plan.new_absences:
        for offset in range(run.days):
            day = run.start + timedelta(days=offset)
            db.execute(
                text(
                    "INSERT OR IGNORE INTO absences (employee_id, date, note, created_at) "
                    "VALUES (:e, :d, :note, :now)"
                ),
                {"e": run.employee_id, "d": day.isoformat(), "note": note, "now": now},
            )

    for leave_id, emp, _ot, _os, _oe, new_type, new_start, new_end in plan.corrections:
        db.execute(
            text("UPDATE leaves SET deleted_at = :now WHERE id = :i"), {"now": now, "i": leave_id}
        )
        if new_type and new_start and new_end:
            days = (date.fromisoformat(new_end) - date.fromisoformat(new_start)).days + 1
            db.execute(
                text(
                    "INSERT INTO leaves (employee_id, leave_type, start_date, end_date, days, "
                    "status, notes, created_at) VALUES (:e, :t, :s, :d, :n, 'Approved', :note, :now)"
                ),
                {
                    "e": emp,
                    "t": new_type,
                    "s": new_start,
                    "d": new_end,
                    "n": days,
                    "note": note,
                    "now": now,
                },
            )

    for leave_id, _emp, start, end in plan.retypes:
        days = (date.fromisoformat(end) - date.fromisoformat(start)).days + 1
        db.execute(
            text(
                "UPDATE leaves SET leave_type = 'Annual Leave', days = :n, updated_at = :now WHERE id = :i"
            ),
            {"n": days, "now": now, "i": leave_id},
        )

    for leave_id, _emp, _s, _e in plan.deletions:
        db.execute(
            text("UPDATE leaves SET deleted_at = :now WHERE id = :i"), {"now": now, "i": leave_id}
        )

    for employee_id, new, _old in plan.end_date_fixes:
        db.execute(
            text("UPDATE employees SET end_date = :d, updated_at = :now WHERE id = :i"),
            {"d": new, "now": now, "i": employee_id},
        )

    if plan.duplicate:
        drop, keep, end = plan.duplicate
        db.execute(
            text(
                "UPDATE employees SET status = 'Resigned', end_date = :d, updated_at = :now "
                "WHERE id = :i"
            ),
            {"d": end, "now": now, "i": keep},
        )
        # Re-point first: foreign_keys=ON, and the dropped record owns books,
        # violations and documents that must survive on the surviving ID.
        # Discovered here rather than read from plan.duplicate_refs: rows
        # inserted earlier in this same run may have landed on `drop`, and the
        # plan-time snapshot would not know about them — the DELETE below would
        # then abort the whole apply on the RESTRICT foreign key.
        for table in employee_tables(db):
            db.execute(
                text(f'UPDATE "{table}" SET employee_id = :keep WHERE employee_id = :drop'),
                {"keep": keep, "drop": drop},
            )
        db.execute(text("DELETE FROM employees WHERE id = :i"), {"i": drop})

    for employee_id, designation_id, _name in plan.designations:
        db.execute(
            text("UPDATE employees SET designation_id = :d, updated_at = :now WHERE id = :i"),
            {"d": designation_id, "now": now, "i": employee_id},
        )

    db.commit()


def verify(series: dict[str, dict[date, str]], db: Session) -> int:
    """Regenerate June and July from the DB and diff every cell. Returns cells."""

    employees, leaves = load_db_state(db)
    absences: dict[str, set[date]] = defaultdict(set)
    for row in db.execute(text("SELECT employee_id, date FROM absences")).all():
        parsed = _as_date(row.date)
        if parsed:
            absences[row.employee_id].add(parsed)

    spans: dict[str, list[LeaveSpan]] = defaultdict(list)
    for row in leaves:
        start, end = _as_date(row["start_date"]), _as_date(row["end_date"])
        if start and end:
            spans[str(row["employee_id"])].append(
                LeaveSpan(str(row["leave_type"]), start, end, str(row["status"]))
            )

    total = 0
    for month in (6, 7):
        days = calendar.monthrange(2026, month)[1]
        month_start, month_end = date(2026, month, 1), date(2026, month, days)
        sheet_rows = {
            employee_id: {day: code for day, code in by_day.items() if day.month == month}
            for employee_id, by_day in series.items()
        }
        diff: Counter[tuple[str, str]] = Counter()
        touched: Counter[str] = Counter()
        compared = 0
        for employee_id, by_day in sheet_rows.items():
            if not by_day or employee_id not in employees:
                continue
            record = employees[employee_id]
            codes = month_codes(
                2026,
                month,
                doj=_as_date(record["doj"]),
                end_date=_as_date(record["end_date"]),
                leaves=spans.get(employee_id, []),
                absences=absences.get(employee_id, set()),
            )
            for day in range(1, days + 1):
                expected = (codes[day - 1] or "").strip().upper()
                actual = by_day.get(date(2026, month, day), "")
                compared += 1
                if expected != actual:
                    diff[(actual, expected)] += 1
                    touched[employee_id] += 1
        cells = sum(diff.values())
        total += cells
        pct = 100 - (100 * cells / compared) if compared else 100.0
        print(
            f"[verify] 2026-{month:02d}: {cells} differing cells of {compared} ({pct:.2f}% match), "
            f"{len(touched)} employees"
        )
        for (actual, expected), count in diff.most_common():
            print(f"           {count:4}  sheet={actual!r:5} engine={expected!r}")
        if touched:
            print(f"           employees: {dict(touched.most_common(10))}")
        roster = {
            employee_id
            for employee_id, record in employees.items()
            if in_roster(
                doj=_as_date(record["doj"]),
                end_date=_as_date(record["end_date"]),
                month_start=month_start,
                month_end=month_end,
            )
        }
        on_sheet = {e for e, by_day in sheet_rows.items() if by_day}
        print(
            f"           roster: db={len(roster)} sheet={len(on_sheet)} "
            f"db-only={sorted(roster - on_sheet)} sheet-only={sorted(on_sheet - roster)}"
        )
    return total


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--db", default=str(DEFAULT_DB), help="SQLite DB path (default: live data/gssg.db)"
    )
    parser.add_argument("--apply", action="store_true", help="mutate (default is dry-run)")
    parser.add_argument(
        "--verify", action="store_true", help="diff June+July against the workbooks and exit"
    )
    args = parser.parse_args()

    db = _session_for(Path(args.db))
    try:
        series = read_series()
        print(f"[import] db={args.db}")
        print(f"[import] workbooks: {len(MONTH_FILES)} months, {len(series)} employees seen")
        if args.verify:
            verify(series, db)
            return
        plan = build_plan(series, db)
        print_plan(plan)
        if not args.apply:
            print("[import] DRY-RUN only. Re-run with --apply after backup + approval.")
            return
        apply_plan(db, plan)
        print("[import] applied.")
        verify(series, db)
    finally:
        db.close()


if __name__ == "__main__":
    main()
