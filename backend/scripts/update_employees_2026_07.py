"""One-off roster sync from the up-to-date project salary workbook.

Source: ``E:\\Al Watbha Shares\\مرتب المشروع الوظيفي.xlsx`` (the live salary
sheet the operations office maintains). The DB layout of this workbook DIFFERS
from the old v3 importer — the ``المرتب`` master sheet now carries **only** the
Arabic name (no English column) and adds the duty group / assigned post:

    col1 ID | col2 name_ar | col3 nationality | col4 phone | col5 uae_id
    col6 position_ar (المسمى الوظيفي) | col7 doj | col8 duty_unit (مكان العمل)
    col9 duty_post (العمل المسند)

Reconciliation rules (agreed with the operator, 2026-07-01):

* Match on the G-number.
* **Excel wins** on matched rows: overwrite contact/nationality/uae_id_no/
  position_ar/doj/duty_unit/duty_post. Refresh name_ar only when it changed.
  **Never touch name_en** (hand-curated transliterations live in the DB).
* New rows (in Excel, not in DB) are inserted; the missing-language name is
  filled from ``NAME_TRANSLATIONS`` below (Arabic→English for Gulf staff,
  English→Arabic for the Nepali support workers). position defaults to the
  Security-Guard pair.
* The lone transfer-sheet employee G0984 is added (duty_unit=النقل).
* The 20 people on the ``الإستقالات`` sheet get status (Terminated/Resigned)
  + end_date + the Arabic reason appended to notes.
* Employees present in the DB but absent from the Excel are LEFT UNTOUCHED.

Run:  python backend/scripts/update_employees_2026_07.py --apply
      (``--dry-run`` prints the report without writing)
"""

from __future__ import annotations

import argparse
import shutil
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

XLSX = Path(r"E:\Al Watbha Shares\مرتب المشروع الوظيفي.xlsx")
DB = Path(__file__).resolve().parents[2] / "data" / "gssg.db"

MASTER_SHEET = "المرتب"
RESIGN_SHEET = "الإستقالات "
TRANSFER_SHEET = "النقل "

# --- name fills for the rows Excel only carries in one language ------------
# Arabic-source new staff -> English transliteration.
# Latin-source support/transfer staff -> Arabic transliteration.
NAME_TRANSLATIONS: dict[str, dict[str, str]] = {
    # Latin -> Arabic (Nepali support group + transfer)
    "G0987": {"name_en": "Dambar Bahadur Shrestha", "name_ar": "دامبار باهادور شريستا"},
    "G1000": {"name_en": "Jiwan Singh Ghatal", "name_ar": "جيوان سينغ غاتال"},
    "G1016": {"name_en": "Phanindra Bahadur Sapkota", "name_ar": "فانيندرا باهادور سابكوتا"},
    "G1019": {"name_en": "Prem Kumar Shrestha", "name_ar": "بريم كومار شريستا"},
    "G1033": {"name_en": "Saroj B K", "name_ar": "ساروج بي كي"},
    "G1107": {"name_en": "Rajan Kumar Karki", "name_ar": "راجان كومار كاركي"},
    "G1111": {"name_en": "Ram Kaji Thapa", "name_ar": "رام كاجي ثابا"},
    "G2051": {"name_en": "Udaya Bahadur B K", "name_ar": "أوداية باهادور بي كي"},
    "G2218": {"name_en": "Hukum Bahadur B K", "name_ar": "حكوم باهادور بي كي"},
    "G0984": {"name_en": "Buddhi Bahadur Gurung", "name_ar": "بودي باهادور غورونغ"},
    # Arabic -> English (Gulf staff)
    "G5130": {"name_ar": "عبدالعزيز محمد سالم سيف المقبالي", "name_en": "Abdulaziz Mohammed Salem Saif Almuqbali"},
    "G5244": {"name_ar": "عادل محمد عادل محمد اسامي", "name_en": "Adel Mohammed Adel Mohammed Asami"},
    "G5260": {"name_ar": "المهدي لمنور", "name_en": "Almahdi Lamnawar"},
    "G5524": {"name_ar": "خليفة بن جمعة بن سعيد الشكري", "name_en": "Khalifa bin Juma bin Saeed Alshukri"},
    "G5527": {"name_ar": "جمعه مهير حمد الرحمة الدرمكي", "name_en": "Juma Muhair Hamad Alrahma Aldarmaki"},
    "G5528": {"name_ar": "مروان عبدالله سعيد محمد المرزوقي", "name_en": "Marwan Abdulla Saeed Mohammed Almarzouqi"},
    "G5530": {"name_ar": "أمين الرطب", "name_en": "Amine Errateb"},
    "G5556": {"name_ar": "علي عبيد علي راشد الزعابي", "name_en": "Ali Obaid Ali Rashed Alzaabi"},
    "G5557": {"name_ar": "سليمان سعيد عبدلله سلطان الشامسي", "name_en": "Suleiman Saeed Abdulla Sultan Alshamsi"},
    "G5558": {"name_ar": "سرور بن عبيد بن خميس المعمري", "name_en": "Suroor bin Obaid bin Khamis Almemari"},
    "G5566": {"name_ar": "عبدالرحمن بن فايل بن حمد بن محمد السنيدي", "name_en": "Abdulrahman bin Fayel bin Hamad bin Mohammed Alsunaidi"},
    "G5677": {"name_ar": "عوض بن راشد بن عبدالله الفليتي", "name_en": "Awadh bin Rashed bin Abdulla Alflaiti"},
}

# مكان العمل normalized to the style already in the DB ("الدوام الرسمي").
DUTY_UNIT_MAP = {
    "دوام رسمي": "الدوام الرسمي",
    "دعم1": "دعم 1",
    "دعم2": "دعم 2",
    "دعم3": "دعم 3",
}


def norm(x: object) -> str:
    return str(x).strip() if x is not None else ""


def has_arabic(s: str) -> bool:
    return any("؀" <= ch <= "ۿ" for ch in s or "")


def parse_date(v: object) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = norm(v)
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def duty_unit(raw: str) -> str:
    return DUTY_UNIT_MAP.get(raw, raw)


def load_master() -> dict[str, dict]:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[MASTER_SHEET]
    out: dict[str, dict] = {}
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if not r or r[1] is None:
            continue
        gid = norm(r[1])
        if not gid or gid in out:
            continue
        out[gid] = {
            "name_ar": norm(r[2]) if len(r) > 2 else "",
            "nationality": norm(r[3]) if len(r) > 3 else "",
            "contact": norm(r[4]) if len(r) > 4 else "",
            "uae_id_no": norm(r[5]) if len(r) > 5 else "",
            "position_ar": norm(r[6]) if len(r) > 6 else "",
            "doj": parse_date(r[7]) if len(r) > 7 else None,
            "duty_unit": duty_unit(norm(r[8])) if len(r) > 8 else "",
            "duty_post": norm(r[9]) if len(r) > 9 else "",
        }
    wb.close()
    return out


def load_resignations() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[RESIGN_SHEET]
    out: list[dict] = []
    for r in list(ws.iter_rows(values_only=True))[3:]:
        if not r or len(r) < 2 or not norm(r[1]).startswith("G"):
            continue
        typ = norm(r[7]) if len(r) > 7 else ""
        status = "Terminated" if "إنهاء" in typ else "Resigned"
        out.append({
            "id": norm(r[1]),
            "status": status,
            "end_date": parse_date(r[6]) if len(r) > 6 else None,
            "reason": norm(r[8]) if len(r) > 8 else "",
            "type_ar": typ,
        })
    wb.close()
    return out


def load_transfer() -> dict | None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[TRANSFER_SHEET]
    out = None
    for r in ws.iter_rows(values_only=True):
        if r and norm(r[0]) == "G0984":
            out = {
                "id": "G0984",
                "nationality": norm(r[2]),
                "contact": norm(r[3]),
                "uae_id_no": norm(r[4]),
                "position_ar": norm(r[5]),
                "doj": parse_date(r[6]),
                "duty_unit": "النقل",
                "duty_post": norm(r[7]),  # transfer destination (الحدود)
            }
            break
    wb.close()
    return out


# fields refreshed on a matched row (Excel wins). name_ar handled separately.
MATCH_FIELDS = ("contact", "nationality", "uae_id_no", "position_ar",
                "doj", "duty_unit", "duty_post")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="write changes")
    ap.add_argument("--dry-run", action="store_true", help="report only")
    args = ap.parse_args()
    write = args.apply and not args.dry_run

    if not XLSX.exists():
        print(f"ERROR: workbook not found: {XLSX}", file=sys.stderr)
        return 2
    if not DB.exists():
        print(f"ERROR: db not found: {DB}", file=sys.stderr)
        return 2

    master = load_master()
    resignations = load_resignations()
    transfer = load_transfer()

    con = sqlite3.connect(str(DB))
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    db_rows = {r["id"]: dict(r) for r in cur.execute("SELECT * FROM employees")}

    # ---- backup (consistent online snapshot, includes WAL) ----------------
    backup_path = None
    if write:
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_path = DB.with_name(f"gssg.db.bak-emp-sync-{ts}")
        bck = sqlite3.connect(str(backup_path))
        con.backup(bck)
        bck.close()

    now = datetime.now().isoformat(sep=" ")  # match stored '%Y-%m-%d %H:%M:%S.%f'
    updates: list[str] = []
    inserts: list[str] = []
    resign_updates: list[str] = []

    # ---- 1. matched updates + new inserts from master ---------------------
    for gid, m in master.items():
        if gid in db_rows:
            row = db_rows[gid]
            changes: dict[str, object] = {}
            for f in MATCH_FIELDS:
                new = m[f]
                if f == "doj":
                    new = new.isoformat() if new else None
                old = row.get(f)
                # only overwrite with a non-empty Excel value
                if new not in (None, "") and norm(old) != norm(new):
                    changes[f] = new
            # name_ar: prefer a curated translation; otherwise refresh from
            # Excel ONLY when the Excel value is actually Arabic script. The
            # master sheet stores the Nepali support workers' names in Latin,
            # so a naive refresh would clobber their curated Arabic name with
            # Latin text (and never converge). name_en is never touched.
            tr = NAME_TRANSLATIONS.get(gid, {})
            desired_ar = tr.get("name_ar") or (
                m["name_ar"] if has_arabic(m["name_ar"]) else None)
            if desired_ar and norm(row.get("name_ar")) != norm(desired_ar):
                changes["name_ar"] = desired_ar
            if changes:
                updates.append(f"{gid}: " + ", ".join(
                    f"{k}={row.get(k)!r}->{v!r}" for k, v in changes.items()))
                if write:
                    changes["updated_at"] = now
                    sets = ", ".join(f"{k}=?" for k in changes)
                    cur.execute(f"UPDATE employees SET {sets} WHERE id=?",
                                (*changes.values(), gid))
        else:
            tr = NAME_TRANSLATIONS.get(gid, {})
            name_ar = tr.get("name_ar") or m["name_ar"]
            name_en = tr.get("name_en") or m["name_ar"]  # fallback: never null
            pos_ar = m["position_ar"] or "حارس أمن"
            pos_en = "Security Guard" if pos_ar == "حارس أمن" else pos_ar
            inserts.append(f"{gid}: {name_en} / {name_ar} | {m['duty_unit']} / {m['duty_post']}")
            if write:
                cur.execute(
                    """INSERT INTO employees
                       (id,name_en,name_ar,nationality,contact,uae_id_no,
                        position,position_ar,doj,duty_unit,duty_post,status,
                        msg_language,created_at,updated_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (gid, name_en, name_ar, m["nationality"] or None,
                     m["contact"] or None, m["uae_id_no"] or None,
                     pos_en, pos_ar,
                     m["doj"].isoformat() if m["doj"] else None,
                     m["duty_unit"] or None, m["duty_post"] or None,
                     "Active", "ar", now, now))

    # ---- 2. transfer-only employee G0984 ----------------------------------
    if transfer and transfer["id"] not in db_rows and transfer["id"] not in master:
        gid = transfer["id"]
        tr = NAME_TRANSLATIONS.get(gid, {})
        pos_ar = transfer["position_ar"] or "حارس أمن"
        pos_en = "Security Guard" if pos_ar == "حارس أمن" else pos_ar
        inserts.append(f"{gid}: {tr.get('name_en')} / {tr.get('name_ar')} | النقل / {transfer['duty_post']}")
        if write:
            cur.execute(
                """INSERT INTO employees
                   (id,name_en,name_ar,nationality,contact,uae_id_no,
                    position,position_ar,doj,duty_unit,duty_post,status,
                    notes,msg_language,created_at,updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (gid, tr.get("name_en"), tr.get("name_ar"),
                 transfer["nationality"] or None, transfer["contact"] or None,
                 transfer["uae_id_no"] or None, pos_en, pos_ar,
                 transfer["doj"].isoformat() if transfer["doj"] else None,
                 "النقل", transfer["duty_post"] or None, "Active",
                 f"نقل إلى {transfer['duty_post']} بتاريخ {transfer['doj']}",
                 "ar", now, now))

    # ---- 3. resignations: status + end_date + reason ----------------------
    for res in resignations:
        gid = res["id"]
        if gid not in db_rows:
            resign_updates.append(f"{gid}: SKIP (not in DB)")
            continue
        row = db_rows[gid]
        end = res["end_date"].isoformat() if res["end_date"] else None
        note = (row.get("notes") or "").strip()
        reason_line = f"{res['type_ar']} — {res['reason']}".strip(" —")
        new_note = (note + ("\n" if note else "") + reason_line) if reason_line and reason_line not in note else note
        resign_updates.append(
            f"{gid}: status={row.get('status')!r}->{res['status']!r} end_date={row.get('end_date')!r}->{end!r} :: {reason_line}")
        if write:
            cur.execute(
                "UPDATE employees SET status=?, end_date=?, notes=?, updated_at=? WHERE id=?",
                (res["status"], end, new_note or None, now, gid))

    if write:
        con.commit()
    con.close()

    # ---- report -----------------------------------------------------------
    print("=" * 72)
    print(f"MODE: {'APPLY (written)' if write else 'DRY-RUN (no writes)'}")
    if backup_path:
        print(f"Backup: {backup_path}")
    print(f"Master rows: {len(master)}  |  DB rows before: {len(db_rows)}")
    print("-" * 72)
    print(f"[UPDATED matched] {len(updates)}")
    for line in updates:
        print("   " + line)
    print(f"\n[INSERTED new] {len(inserts)}")
    for line in inserts:
        print("   " + line)
    print(f"\n[RESIGNATIONS] {len(resign_updates)}")
    for line in resign_updates:
        print("   " + line)
    print("=" * 72)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
