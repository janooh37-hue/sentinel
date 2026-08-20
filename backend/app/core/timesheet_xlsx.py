"""The two monthly deliverables, rendered onto the paper already in circulation.

One template, two variants. ``attendance`` is HR's sheet: column E prints the
English designation and the day cells print :attr:`GridRow.codes`, the truth
about each day. ``statistics`` is the client's: column E prints the Arabic
designation, the day cells print :attr:`GridRow.stat_codes`, and the roster is
split into the contracted posts and the surplus above them, separated by two
empty rows while column A keeps counting straight through.

Nothing here reads a database. A :class:`MonthGrid` plus the committed template
is the whole input, so the same grid always renders the same workbook.

**Five footer formulas deliberately diverge from the June source, because June is
buggy.** June counts OFF as ``=COUNTIF(F7:AJ288,"OFF")`` — off by one row at both
ends — and writes ``-``, ``R`` and ``S`` as ``=COUNTIF(F6:AJ287,D300)`` style cell
references, with ``R`` and ``S`` stopping 24 rows short of the roster. All four are
normalised here. Consequence, stated plainly because someone will compare: the
OFF, R and S totals in a generated workbook will **not** match a hand file's for
the same month, and it is the generated one that is right. Do not "fix" these
back to match June. The fifth is whitespace with no numeric effect: June's NG row
is ``=COUNTIF(F6:AJ287, "NG")`` with a stray space after the comma, and the
generated form has none — recorded so that someone diffing a generated file
against a hand file can account for every textual difference.

**The code data validation deliberately changes shape — its formula source, and
nothing else.** The source's three list validations point at ``$D$295:$D$304``, a
reference into its own footer code rows that only resolves while the footer sits
exactly there. The renderer emits one validation carrying the literal list
instead, which openpyxl needs quoted. Both message flags stay on, as the source
carries them: openpyxl defaults them off, and with ``showErrorMessage`` off Excel
shows the dropdown but silently accepts anything typed over it, which would drop
the input guard the paper has today.

Two openpyxl traps, both measured against the June workbook:

* The source stores its conditional-format fills on ``bgColor`` with no
  ``patternType``. ``PatternFill(start_color=..., fill_type="solid")`` sets
  ``fgColor`` and produces different XML, so the rules are built as
  ``DifferentialStyle(fill=PatternFill(bgColor=...))``. The colours carry their
  ``FF`` alpha byte because a bare six-digit hex is stored as ``00``-alpha.
* The source's sick-leave rule tests the literal ``"SL"`` *without* the trailing
  space, even though ``AO5`` and :data:`CODE_SICK` are ``"SL "`` *with* it — a
  latent bug that means the rule never fires on the hand file. The rebuilt rule
  uses ``"SL "``.

``_parts`` carries **two** specimen data rows because the source's first data row
is not typical: June row 6 draws a ``medium`` top border across AK-AP, the rule
that separates the day-number header from the body, and June's other 281 rows do
not. Output row 6 therefore takes specimen 1 and every row after it takes
specimen 2 — a positional rule, so the header rule draws exactly once whatever
occupies row 6: a data row, a statistics gap row, or the blank row an empty
roster gets.
"""

from __future__ import annotations

import io
from collections.abc import Mapping, Sequence
from copy import copy, deepcopy
from pathlib import Path
from typing import Any, Final, NamedTuple

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.api.errors import NotFoundError, ValidationFailedError
from app.config import get_settings
from app.core.constants import ARABIC_MONTHS
from app.core.timesheet_codes import (
    CODE_ABSENT,
    CODE_ANNUAL,
    CODE_NATIONAL,
    CODE_NEW,
    CODE_SICK,
)
from app.services.timesheet_service import CODE_BLOCKED, GridRow, MonthGrid

# --------------------------------------------------------------------------- #
# The paper's geometry
# --------------------------------------------------------------------------- #

_TEMPLATE_NAME: Final[str] = "GSSG-HR_Monthly_Time_Sheet.xlsx"

#: The visible sheet, and the hidden sheet that carries the two specimen data
#: rows and the footer block. ``_parts`` is deleted last, once all of it is copied.
_SHEET: Final[str] = "Sheet1"
_PARTS: Final[str] = "_parts"

#: ``_parts`` row 1 models output row 6, which carries the rule under the
#: day-number header; row 2 models every row after it. Rows 3-21 are the footer.
_SPECIMEN_ROWS: Final[tuple[int, int]] = (1, 2)
_FOOTER_ROWS: Final[range] = range(3, 22)

_FIRST_DATA_ROW: Final[int] = 6
_FIRST_DAY_COLUMN: Final[int] = 6  # F = day 1
_LAST_COLUMN: Final[int] = 42  # AP
_TOTAL_COLUMNS: Final[range] = range(37, 43)  # AK..AP, the per-row totals
_CODE_COLUMN: Final[int] = 4  # D, in the footer's code rows
_VALUE_COLUMN: Final[int] = 5  # E, ditto
_GAP_ROWS: Final[int] = 2

#: Uppercase English month abbreviations, January at index 0. Spelled out rather
#: than taken from ``calendar.month_abbr``, which goes through ``strftime`` and
#: therefore through the process locale — the sheet title and the span's sheet
#: names must read the same on every machine.
_MONTH_ABBR: Final[tuple[str, ...]] = (
    "JAN",
    "FEB",
    "MAR",
    "APR",
    "MAY",
    "JUN",
    "JUL",
    "AUG",
    "SEP",
    "OCT",
    "NOV",
    "DEC",
)

VARIANT_ATTENDANCE: Final[str] = "attendance"
VARIANT_STATISTICS: Final[str] = "statistics"
VARIANTS: Final[frozenset[str]] = frozenset({VARIANT_ATTENDANCE, VARIANT_STATISTICS})

_DRIVERS_SHEET: Final[str] = "drivers"

#: File-name stems, exactly as the share already spells them.
_ATTENDANCE_STEM: Final[str] = "كشف حضور شهر"
_STATISTICS_STEM: Final[str] = "الاحصائية شهر"
_SINGLE_STEM: Final[str] = "كشف حضور"
_DRIVERS_SUFFIX: Final[str] = "للسائقين"

#: The six per-row totals, verbatim from the source. ``AN`` really does span
#: ``AK``, and ``AO`` really does reference ``$AO$5`` instead of a literal —
#: ``AO5`` holds ``"SL "``, trailing space included.
_ROW_FORMULAS: Final[tuple[tuple[int, str], ...]] = (
    (37, '=COUNTIF(F{row}:AJ{row},"P")'),  # AK — Total day
    (38, '=COUNTIF(F{row}:AJ{row},"OFF")'),  # AL — Off
    (39, '=COUNTIF(F{row}:AJ{row},"AB")'),  # AM
    (40, '=COUNTIF(F{row}:AK{row},"AL")'),  # AN — spans AK in the original
    (41, "=COUNTIF(F{row}:AJ{row},$AO$5)"),  # AO — $AO$5 holds "SL "
    (42, '=COUNTIF(F{row}:AJ{row},"TR")'),  # AP
)

#: Footer code → the per-row total column that already counts it. Anything else
#: in the code column is counted straight off the grid with a COUNTIF.
_SUMMARY_COLUMN: Final[Mapping[str, str]] = {
    CODE_SICK: "AO",
    CODE_ANNUAL: "AN",
    CODE_ABSENT: "AM",
    CODE_NATIONAL: "AP",
    "P": "AK",
    "OFF": "AL",  # normalised: June counts this one row off at both ends
}

#: (code, fill, font) for the cell rules, from UI spec §3.2 and §15. The leading
#: ``FF`` is the alpha byte the source's dxf fills carry; a six-digit hex would
#: be stored as ``00``-alpha and diverge from the paper.
_CELL_RULES: Final[tuple[tuple[str, str, str | None], ...]] = (
    (CODE_ANNUAL, "FFBDD7EE", None),
    (CODE_SICK, "FFC6E0B4", None),  # "SL " — the source's rule drops the space
    (CODE_ABSENT, "FFFFC7CE", "FF9C0006"),
    (CODE_NATIONAL, "FFCC99FF", None),
    (CODE_NEW, "FFFF9900", None),
    (CODE_BLOCKED, "FF990033", "FFFFFFFF"),  # the design lock's red block
)

#: The codes an operator may type into a day cell, as a literal list. Quoted
#: because that is how openpyxl passes an inline list to Excel.
_CODE_LIST: Final[str] = '"P,AL,SL ,AB,TR,NG,-,X"'


# --------------------------------------------------------------------------- #
# Loading
# --------------------------------------------------------------------------- #


def _template_path() -> Path:
    """The committed template, resolved **now**.

    Never at import time: the PyInstaller build moves ``templates_dir`` to
    ``sys._MEIPASS/templates/``, and ``GSSG_TEMPLATES_DIR`` may move it again.
    """

    return get_settings().templates_dir / _TEMPLATE_NAME


def _load() -> tuple[Any, Any, Any]:
    workbook = load_workbook(_template_path())
    return workbook, workbook[_SHEET], workbook[_PARTS]


def _finish(workbook: Any) -> bytes:
    """Drop the helper sheet — last, once every style has been copied off it."""

    del workbook[_PARTS]
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _sheets(workbook: Any, template_sheet: Any, count: int) -> list[Any]:
    """``count`` copies of the pristine header sheet, one per grid.

    ``copy_worksheet`` carries cells, styles, merges and dimensions but drops the
    logo and the frozen panes, and an :class:`~openpyxl.drawing.image.Image`
    cannot be shared between sheets — writing one consumes its stream — so each
    copy gets its own image built from the same bytes.
    """

    sheets = [template_sheet]
    for _ in range(count - 1):
        extra = workbook.copy_worksheet(template_sheet)
        for source in template_sheet._images:
            image = XLImage(io.BytesIO(source.ref.getvalue()))
            image.anchor = deepcopy(source.anchor)
            extra.add_image(image)
        extra.freeze_panes = template_sheet.freeze_panes
        sheets.append(extra)
    return sheets


# --------------------------------------------------------------------------- #
# The data band
# --------------------------------------------------------------------------- #


class _Specimen(NamedTuple):
    """One styled model row read off ``_parts``: 42 cell styles and a height."""

    styles: tuple[Any, ...]
    height: float | None


def _specimens(parts: Any) -> tuple[_Specimen, _Specimen]:
    """The first-data-row model and the subsequent-row model.

    Both heights are read, never hardcoded, so the template and the renderer
    cannot drift apart.
    """

    def read(row: int) -> _Specimen:
        return _Specimen(
            tuple(parts.cell(row, column)._style for column in range(1, _LAST_COLUMN + 1)),
            parts.row_dimensions[row].height,
        )

    return read(_SPECIMEN_ROWS[0]), read(_SPECIMEN_ROWS[1])


def _style_row(sheet: Any, specimens: tuple[_Specimen, _Specimen], row: int) -> None:
    """Style one output row from the model that belongs at that position.

    Positional, not per-caller: output row 6 sits under the day-number header and
    carries the ``medium`` rule that separates it, and every row after it must not
    repeat that rule. So the header rule draws exactly once whatever occupies row
    6 — a data row, a statistics gap row, or an empty roster's blank row.
    """

    specimen = specimens[0] if row == _FIRST_DATA_ROW else specimens[1]
    sheet.row_dimensions[row].height = specimen.height
    for column, style in enumerate(specimen.styles, start=1):
        sheet.cell(row, column)._style = copy(style)


def _write_row(sheet: Any, row: int, entry: GridRow, *, statistics: bool, days: int) -> None:
    sheet.cell(row, 1, entry.row_no)
    sheet.cell(row, 2, entry.employee_id)
    sheet.cell(row, 3, entry.name_en)
    sheet.cell(row, 4, entry.nationality_en)
    sheet.cell(row, 5, entry.designation_ar if statistics else entry.designation_en)

    codes = entry.stat_codes if statistics else entry.codes
    for day, code in enumerate(codes[:days], start=1):
        if code is not None:
            sheet.cell(row, _FIRST_DAY_COLUMN + day - 1, code)

    for column, formula in _ROW_FORMULAS:
        sheet.cell(row, column, formula.format(row=row))


def _write_band(
    sheet: Any, parts: Any, grid: MonthGrid, rows: Sequence[GridRow], *, statistics: bool
) -> int:
    """Write the data rows and return the last one.

    The statistics variant separates the surplus block from the contracted posts
    with two empty rows. They are emitted before the first block-2 row, which is
    row 6 itself when the contracted post count is zero and block 1 is therefore
    empty.
    """

    specimens = _specimens(parts)
    target = _FIRST_DATA_ROW
    pending_gap = statistics

    for entry in rows:
        if pending_gap and entry.stat_block != 1:
            for _ in range(_GAP_ROWS):
                _style_row(sheet, specimens, target)
                target += 1
            pending_gap = False
        _style_row(sheet, specimens, target)
        _write_row(sheet, target, entry, statistics=statistics, days=grid.days_in_month)
        target += 1

    if target == _FIRST_DATA_ROW:
        # Nobody on this roster. The band still gets one row, because the footer
        # ranges have to be well formed: `=SUM(AK6:AK5)` is a reversed reference
        # that folds the day-number header row into the total.
        _style_row(sheet, specimens, target)
        target += 1
    return target - 1


# --------------------------------------------------------------------------- #
# The footer
# --------------------------------------------------------------------------- #


def _copy_footer(sheet: Any, parts: Any, last_row: int) -> None:
    for offset, source_row in enumerate(_FOOTER_ROWS):
        target_row = last_row + 1 + offset
        height = parts.row_dimensions[source_row].height
        if height is not None:
            sheet.row_dimensions[target_row].height = height
        for column in range(1, _LAST_COLUMN + 1):
            source = parts.cell(source_row, column)
            if source.has_style:
                sheet.cell(target_row, column)._style = copy(source._style)
            if source.value is not None:
                sheet.cell(target_row, column, source.value)


def _merge(sheet: Any, row: int, first: int, last: int, *, through: int | None = None) -> None:
    sheet.merge_cells(start_row=row, start_column=first, end_row=through or row, end_column=last)


def _write_footer(sheet: Any, parts: Any, last_row: int) -> None:
    """The 19-row block below the roster: legend, signatures, totals, code table."""

    _copy_footer(sheet, parts, last_row)

    legend = last_row + 1
    signatures = last_row + 2
    totals = last_row + 3
    header = last_row + 7
    first_code = last_row + 8
    last_code = last_row + 18
    total_days = last_row + 19

    _merge(sheet, legend, 1, _LAST_COLUMN)  # A:AP
    _merge(sheet, signatures, 1, 13)  # A:M
    _merge(sheet, signatures, 14, 29)  # N:AC
    _merge(sheet, signatures, 30, _LAST_COLUMN)  # AD:AP
    _merge(sheet, header, 1, 2)  # A:B on the S.no row
    _merge(sheet, first_code, 1, 2, through=last_code)  # A:B down the eleven codes
    _merge(sheet, total_days, 1, 4)  # A:D on Total Days

    for column in _TOTAL_COLUMNS:
        letter = get_column_letter(column)
        sheet.cell(totals, column, f"=SUM({letter}{_FIRST_DATA_ROW}:{letter}{last_row})")

    for row in range(first_code, last_code + 1):
        code = sheet.cell(row, _CODE_COLUMN).value
        if code is None:
            continue
        summary = _SUMMARY_COLUMN.get(code)
        if summary is None:
            value = f'=COUNTIF(F{_FIRST_DATA_ROW}:AJ{last_row},"{code}")'
        else:
            value = f"={summary}{totals}"
        sheet.cell(row, _VALUE_COLUMN, value)

    sheet.cell(total_days, _VALUE_COLUMN, f"=SUM(E{first_code}:E{last_code})")


# --------------------------------------------------------------------------- #
# Colour and validation, neither of which the template carries
# --------------------------------------------------------------------------- #


def _write_rules(sheet: Any, last_row: int) -> None:
    band = f"F{_FIRST_DATA_ROW}:AJ{last_row}"

    for code, fill, font in _CELL_RULES:
        style = DifferentialStyle(
            font=None if font is None else Font(color=font),
            fill=PatternFill(bgColor=fill),
        )
        sheet.conditional_formatting.add(
            band,
            Rule(type="cellIs", operator="equal", formula=[f'"{code}"'], dxf=style),
        )

    # Both message flags on, as all three of the source's validations carry them:
    # openpyxl defaults them off, and with `showErrorMessage` off Excel renders the
    # dropdown but silently accepts anything typed over it.
    validation = DataValidation(
        type="list",
        formula1=_CODE_LIST,
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
    )
    sheet.add_data_validation(validation)
    validation.add(band)


# --------------------------------------------------------------------------- #
# Public
# --------------------------------------------------------------------------- #


def _require_variant(variant: str) -> None:
    if variant not in VARIANTS:
        raise ValidationFailedError(
            "INVALID_TIMESHEET_VARIANT",
            f"variant {variant!r} is not a recognised time-sheet variant",
            valid=sorted(VARIANTS),
        )


def _month(grid: MonthGrid) -> str:
    return ARABIC_MONTHS[grid.month - 1]


def _row_of(grid: MonthGrid, employee_id: str) -> GridRow:
    for entry in grid.rows:
        if entry.employee_id == employee_id:
            return entry
    raise NotFoundError(
        "EMPLOYEE_NOT_ON_SHEET",
        f"{employee_id!r} is not on the {grid.month}/{grid.year} {grid.sheet} sheet",
        employee_id=employee_id,
    )


def _fill(
    sheet: Any, parts: Any, grid: MonthGrid, rows: Sequence[GridRow], *, statistics: bool
) -> None:
    sheet["D4"] = f"For the Month of :{_MONTH_ABBR[grid.month - 1]}-{grid.year}"
    last_row = _write_band(sheet, parts, grid, rows, statistics=statistics)
    _write_footer(sheet, parts, last_row)
    _write_rules(sheet, last_row)


def render(grid: MonthGrid, *, variant: str = VARIANT_ATTENDANCE) -> bytes:
    """One month of one roster as workbook bytes."""

    _require_variant(variant)
    workbook, sheet, parts = _load()
    _fill(sheet, parts, grid, grid.rows, statistics=variant == VARIANT_STATISTICS)
    return _finish(workbook)


def render_single(grid: MonthGrid, employee_id: str) -> bytes:
    """One employee's attendance row on its own sheet.

    Column A keeps his position on the month's roster rather than renumbering to
    1: the extract is read next to the full sheet, and the position is how HR
    finds him on it.
    """

    workbook, sheet, parts = _load()
    _fill(sheet, parts, grid, [_row_of(grid, employee_id)], statistics=False)
    return _finish(workbook)


def render_single_span(grids: Sequence[MonthGrid], employee_id: str) -> bytes:
    """One employee, one sheet per grid, named by the English month abbreviation.

    The resignation and termination handover: the month of departure and the one
    before it, in the order given. A month he was not on the roster for is his
    sheet with no data row rather than a failure — he may have joined mid-span —
    but a span that never finds him at all is a miss.
    """

    if not grids:
        raise ValueError("a span needs at least one month")

    workbook, template_sheet, parts = _load()
    found = False
    for sheet, grid in zip(_sheets(workbook, template_sheet, len(grids)), grids, strict=True):
        sheet.title = _MONTH_ABBR[grid.month - 1]
        rows = [entry for entry in grid.rows if entry.employee_id == employee_id]
        found = found or bool(rows)
        _fill(sheet, parts, grid, rows, statistics=False)
    if not found:
        raise NotFoundError(
            "EMPLOYEE_NOT_ON_SHEET",
            f"{employee_id!r} is not on any month of the span",
            employee_id=employee_id,
        )
    return _finish(workbook)


def filename_for(grid: MonthGrid, *, variant: str = VARIANT_ATTENDANCE) -> str:
    """The Arabic file name the share already uses for this deliverable."""

    _require_variant(variant)
    stem = _STATISTICS_STEM if variant == VARIANT_STATISTICS else _ATTENDANCE_STEM
    words = [stem, _month(grid)]
    if grid.sheet == _DRIVERS_SHEET:
        words.append(_DRIVERS_SUFFIX)
    return " ".join(words) + ".xlsx"


def filename_for_single(grid: MonthGrid, employee_id: str) -> str:
    """The single-employee extract's name: the man, not the roster, so no
    drivers suffix — it disambiguates two roster workbooks in one folder."""

    return f"{_SINGLE_STEM} {_row_of(grid, employee_id).name_en} {_month(grid)}.xlsx"
