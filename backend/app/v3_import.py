"""v3.5.4 → v4 data migration.

Reads the heterogeneous v3 data directory (one ``.xlsx`` + several ``.json``
files plus a per-employee filesystem hierarchy) and writes a clean
``gssg.db``. The script-level CLI in ``scripts/import_v3.py`` is a thin
wrapper around :func:`run_import`.

Design constraints (from plans/02-data-layer.md):

* **Backups first.** Every read of v3 data is preceded by a snapshot to
  ``<dest_dir>/backups/v3.5.4-<timestamp>/`` so the originals survive any
  catastrophe in the importer.
* **Defensive parsing.** v3 JSON files are unschema'd in practice — rows
  missing keys are logged and skipped rather than raising.
* **Idempotent by refusal.** Running twice without ``--force`` against a
  DB that already has data is rejected. ``--force`` truncates the parity
  tables and re-imports cleanly.
* **No file moves.** The vault index records paths; it does not relocate
  files. Phase 09 owns the move-into-vault flow.
"""

from __future__ import annotations

import json
import logging
import shutil
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import Engine, create_engine, delete, select
from sqlalchemy.orm import Session, sessionmaker

from app.core.constants import DEFAULT_CATEGORIES, EMPLOYEE_STATUS_ACTIVE
from app.db.base import Base
from app.db.models import (
    REF_SEQUENCE_ID,
    AppSetting,
    Book,
    BookCategory,
    BookRefSequence,
    Employee,
    Leave,
    Manager,
    VaultFile,
    Violation,
)
from app.db.session import attach_sqlite_pragmas

logger = logging.getLogger(__name__)


# ----------------------------------------------------------------------------
# Summary & result types
# ----------------------------------------------------------------------------


@dataclass(slots=True)
class ImportSummary:
    """Counts of rows written (or rows that *would* be written under ``--dry``)."""

    employees: int = 0
    leaves: int = 0
    violations: int = 0
    managers: int = 0
    book_categories: int = 0
    books: int = 0
    settings: int = 0
    vault_files: int = 0
    next_ref_number: int = 1
    skipped: list[str] = field(default_factory=list)
    backup_dir: Path | None = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "employees": self.employees,
            "leaves": self.leaves,
            "violations": self.violations,
            "managers": self.managers,
            "book_categories": self.book_categories,
            "books": self.books,
            "settings": self.settings,
            "vault_files": self.vault_files,
            "next_ref_number": self.next_ref_number,
            "skipped": list(self.skipped),
            "backup_dir": str(self.backup_dir) if self.backup_dir else None,
        }


# ----------------------------------------------------------------------------
# Date / value coercion helpers
# ----------------------------------------------------------------------------


def _parse_dmy(value: Any) -> date | None:
    """Parse a ``DD/MM/YYYY`` string (the v3 leave/violation format)."""
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    logger.warning("could not parse date %r", value)
    return None


def _parse_xlsx_date(value: Any) -> date | None:
    """Coerce an xlsx cell value to a date.

    openpyxl emits ``datetime`` for date-typed cells but plain ``str`` for
    cells typed as text — both appear in the live ``employees.xlsx``.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return _parse_dmy(value)


def _parse_book_created_at(date_str: Any, time_str: Any) -> datetime:
    """Combine v3's split ``date`` (YYYY-MM-DD) + ``time`` (HH:MM:SS) fields."""
    d = _parse_dmy(date_str) or datetime.now(UTC).replace(tzinfo=None).date()
    t_str = str(time_str or "00:00:00").strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            t = datetime.strptime(t_str, fmt).time()
            return datetime.combine(d, t)
        except ValueError:
            continue
    return datetime.combine(d, datetime.min.time())


def _coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _ref_suffix_int(ref_number: str) -> int | None:
    """Extract the numeric suffix from ``<cat>-<n>``."""
    if "-" not in ref_number:
        return None
    suffix = ref_number.rsplit("-", 1)[1]
    return int(suffix) if suffix.isdigit() else None


# ----------------------------------------------------------------------------
# Backup
# ----------------------------------------------------------------------------


_V3_FILES: tuple[str, ...] = (
    "employees.xlsx",
    "books_database.json",
    "leave_history.json",
    "violations.json",
    "managers.json",
    "settings.json",
)


def create_backup(src_dir: Path, dest_root: Path) -> Path:
    """Copy every v3 source file into ``<dest_root>/v3.5.4-<ts>/`` and return the dir.

    If the same timestamp already exists (tests re-running inside one second),
    a numeric suffix is appended so backups never collide.
    """
    ts = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    base = dest_root / f"v3.5.4-{ts}"
    backup_dir = base
    suffix = 1
    while backup_dir.exists():
        backup_dir = dest_root / f"v3.5.4-{ts}-{suffix}"
        suffix += 1
    backup_dir.mkdir(parents=True, exist_ok=False)
    for name in _V3_FILES:
        src = src_dir / name
        if src.exists():
            shutil.copy2(src, backup_dir / name)
    logger.info("v3 backup written to %s", backup_dir)
    return backup_dir


# ----------------------------------------------------------------------------
# Per-file parsers — each returns plain dicts that the writer turns into rows
# ----------------------------------------------------------------------------


_EMPLOYEE_COLUMNS: tuple[tuple[int, str], ...] = (
    (1, "id"),           # ID  (G-number)
    (2, "name_en"),      # English name
    (3, "name_ar"),      # Arabic name
    (4, "nationality"),  # الجنسية
    (5, "contact"),      # رقم الهاتف
    (6, "uae_id_no"),    # رقم الهوية
    (7, "position"),     # Tasks (English)
    (8, "position_ar"),  # المهام
    (9, "other"),        # آخرى
    (10, "doj"),         # تاريخ الالتحاق
    (11, "notes"),       # ملاحظات
    (13, "status"),      # الحالة (lazy column 13 in v3 — index 13 here too)
    (14, "end_date"),    # تاريخ نهاية الخدمة
)


def parse_employees(xlsx_path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    """Read the ``'المرتب'`` sheet. Returns ``(rows, skipped_reasons)``."""
    skipped: list[str] = []
    if not xlsx_path.exists():
        skipped.append(f"employees.xlsx missing at {xlsx_path}")
        return [], skipped

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = "المرتب" if "المرتب" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # row indices are 0-based; column 1 is ID
        if not row or row[1] is None:
            continue
        emp: dict[str, Any] = {}
        for col_idx, attr in _EMPLOYEE_COLUMNS:
            if col_idx < len(row):
                emp[attr] = row[col_idx]
        emp_id = _coerce_str(emp.get("id"))
        if not emp_id:
            skipped.append(f"row {row_idx}: blank ID")
            continue
        if not _coerce_str(emp.get("name_en")):
            skipped.append(f"row {row_idx} ({emp_id}): blank name_en")
            continue
        emp["id"] = emp_id
        emp["name_en"] = _coerce_str(emp["name_en"]) or ""
        for k in ("name_ar", "nationality", "contact", "uae_id_no",
                 "position", "position_ar", "other", "notes"):
            emp[k] = _coerce_str(emp.get(k))
        emp["doj"] = _parse_xlsx_date(emp.get("doj"))
        emp["end_date"] = _parse_xlsx_date(emp.get("end_date"))
        emp["status"] = _coerce_str(emp.get("status")) or EMPLOYEE_STATUS_ACTIVE.split(" -")[0]
        rows.append(emp)
    wb.close()
    return rows, skipped


def parse_managers(json_path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    rows, skipped = _read_json_list(json_path)
    out: list[dict[str, Any]] = []
    for i, r in enumerate(rows):
        if not r.get("name_en") and not r.get("name_ar"):
            skipped.append(f"managers[{i}]: blank name")
            continue
        out.append(
            {
                "id": r.get("id"),
                "name_en": _coerce_str(r.get("name_en")),
                "name_ar": _coerce_str(r.get("name_ar")),
                "title": _coerce_str(r.get("title")),
                "sig_path": _coerce_str(r.get("sig_path")),
                "active": bool(r.get("active", True)),
            }
        )
    return out, skipped


def parse_leaves(json_path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    rows, skipped = _read_json_list(json_path)
    out: list[dict[str, Any]] = []
    for i, r in enumerate(rows):
        g = _coerce_str(r.get("g_number"))
        if not g:
            skipped.append(f"leaves[{i}]: blank g_number")
            continue
        start = _parse_dmy(r.get("start_date"))
        end = _parse_dmy(r.get("end_date"))
        if not start or not end:
            skipped.append(f"leaves[{i}] ({g}): missing/unparseable dates")
            continue
        out.append(
            {
                "employee_id": g,
                "leave_type": _coerce_str(r.get("leave_type")) or "Unknown",
                "start_date": start,
                "end_date": end,
                "days": int(r.get("total_days") or 0),
                "status": _coerce_str(r.get("status")) or "Approved",
                "request_date": _parse_dmy(r.get("request_date")),
                "doc_path": _coerce_str(r.get("doc_path")),
                "certificate_path": _coerce_str(r.get("certificate_path")),
                "created_at": _parse_created_at(r.get("created_at")),
            }
        )
    return out, skipped


def parse_violations(json_path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    rows, skipped = _read_json_list(json_path)
    out: list[dict[str, Any]] = []
    for i, r in enumerate(rows):
        g = _coerce_str(r.get("g_number"))
        if not g:
            skipped.append(f"violations[{i}]: blank g_number")
            continue
        d = _parse_dmy(r.get("date"))
        if not d:
            skipped.append(f"violations[{i}] ({g}): unparseable date")
            continue
        out.append(
            {
                "employee_id": g,
                "violation_type": _coerce_str(r.get("violation_type")) or "Violation",
                "date": d,
                "description": _coerce_str(r.get("description")),
                "action_taken": _coerce_str(r.get("action_taken")),
                "deduction_days": int(r.get("deduction_days") or 0),
                "status": _coerce_str(r.get("status")) or "Open",
                "doc_path": _coerce_str(r.get("file_path")),
                "created_at": _parse_created_at(r.get("created_at")),
            }
        )
    return out, skipped


def parse_books(json_path: Path) -> tuple[dict[str, Any], list[str]]:
    """Returns a payload dict ``{categories, books, next_ref_number}`` + skipped."""
    payload, skipped = _read_json_dict(json_path)
    raw_books = payload.get("books", []) if isinstance(payload, dict) else []
    raw_categories = payload.get("categories", {}) if isinstance(payload, dict) else {}
    next_ref = int(payload.get("next_ref_number", 1)) if isinstance(payload, dict) else 1

    categories: dict[str, dict[str, str | None]] = {}
    # Seed defaults so the table always carries the canonical 12 entries.
    for cid, label in DEFAULT_CATEGORIES.items():
        name_en, _, name_ar = label.partition(" - ")
        categories[cid] = {"name_en": name_en, "name_ar": name_ar or None, "prefix": cid}
    # Overlay anything in the file (e.g. updated labels).
    if isinstance(raw_categories, dict):
        for cid, label in raw_categories.items():
            cid = str(cid)
            label = str(label)
            name_en, _, name_ar = label.partition(" - ")
            categories[cid] = {"name_en": name_en, "name_ar": name_ar or None, "prefix": cid}

    books: list[dict[str, Any]] = []
    for i, r in enumerate(raw_books):
        if not isinstance(r, dict):
            skipped.append(f"books[{i}]: not an object")
            continue
        ref = _coerce_str(r.get("ref_number"))
        if not ref:
            skipped.append(f"books[{i}]: blank ref_number")
            continue
        cat_id = _coerce_str(r.get("category")) or ref.split("-", 1)[0]
        if cat_id not in categories:
            categories[cat_id] = {
                "name_en": _coerce_str(r.get("category_name")),
                "name_ar": None,
                "prefix": cat_id,
            }
        emp_g = _coerce_str(r.get("employee_g_number"))
        books.append(
            {
                "ref_number": ref,
                "category_id": cat_id,
                "subject": _coerce_str(r.get("subject")),
                "employee_id": emp_g,
                "employee_name_snapshot": _coerce_str(r.get("employee_name")),
                "notes": _coerce_str(r.get("notes")),
                "doc_path": _coerce_str(r.get("file_path")),
                "created_at": _parse_book_created_at(r.get("date"), r.get("time")),
            }
        )

    # Defensive: ensure next_ref_number is at least one past the highest seen ref.
    max_seen = max(
        (n for n in (_ref_suffix_int(b["ref_number"]) for b in books) if n is not None),
        default=0,
    )
    if next_ref <= max_seen:
        next_ref = max_seen + 1

    return {
        "categories": categories,
        "books": books,
        "next_ref_number": next_ref,
    }, skipped


def parse_settings(json_path: Path) -> tuple[dict[str, str], list[str]]:
    """Settings is a free-form dict; we store each top-level key as a row."""
    payload, skipped = _read_json_dict(json_path)
    out: dict[str, str] = {}
    if isinstance(payload, dict):
        for k, v in payload.items():
            out[str(k)] = json.dumps(v, ensure_ascii=False)
    return out, skipped


def walk_vault_files(
    employee_files_dir: Path, known_employees: set[str]
) -> tuple[list[dict[str, Any]], list[str]]:
    """Index files under ``employee_files/<G>/**`` without moving anything."""
    skipped: list[str] = []
    if not employee_files_dir.exists():
        return [], [f"employee_files dir missing at {employee_files_dir}"]

    out: list[dict[str, Any]] = []
    for g_dir in sorted(p for p in employee_files_dir.iterdir() if p.is_dir()):
        g = g_dir.name
        if g not in known_employees:
            skipped.append(f"vault: unknown employee dir {g}")
            continue
        for sub in g_dir.iterdir():
            if not sub.is_dir():
                continue
            kind = sub.name
            for f in sub.rglob("*"):
                if not f.is_file():
                    continue
                try:
                    size = f.stat().st_size
                except OSError:
                    size = None
                out.append(
                    {
                        "employee_id": g,
                        "kind": kind,
                        "filename": f.name,
                        "path": str(f.relative_to(employee_files_dir.parent)).replace("\\", "/"),
                        "size_bytes": size,
                    }
                )
    return out, skipped


# ----------------------------------------------------------------------------
# JSON helpers
# ----------------------------------------------------------------------------


def _read_json_list(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    if not path.exists():
        return [], [f"{path.name} missing at {path}"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        return [], [f"{path.name}: invalid JSON ({e})"]
    if not isinstance(data, list):
        return [], [f"{path.name}: expected a JSON array, got {type(data).__name__}"]
    return [r for r in data if isinstance(r, dict)], []


def _read_json_dict(path: Path) -> tuple[dict[str, Any], list[str]]:
    if not path.exists():
        return {}, [f"{path.name} missing at {path}"]
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        return {}, [f"{path.name}: invalid JSON ({e})"]
    if not isinstance(data, dict):
        return {}, [f"{path.name}: expected a JSON object, got {type(data).__name__}"]
    return data, []


def _parse_created_at(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    s = _coerce_str(value)
    if s:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return datetime.now(UTC).replace(tzinfo=None)


# ----------------------------------------------------------------------------
# Writer — turns parsed dicts into ORM rows
# ----------------------------------------------------------------------------


# Tables that get wiped under ``--force`` (order matters — children first).
_TRUNCATE_ORDER: tuple[type[Base], ...] = (
    VaultFile,
    AppSetting,
    Book,
    BookRefSequence,
    BookCategory,
    Manager,
    Violation,
    Leave,
    Employee,
)


def _db_is_empty(session: Session) -> bool:
    """True if every parity table has zero rows."""
    return all(
        session.execute(select(model).limit(1)).first() is None
        for model in _TRUNCATE_ORDER
    )


def _truncate(session: Session) -> None:
    for model in _TRUNCATE_ORDER:
        session.execute(delete(model))
    session.flush()


def _write_employees(session: Session, rows: Iterable[dict[str, Any]]) -> int:
    count = 0
    for r in rows:
        session.add(Employee(**r))
        count += 1
    session.flush()
    return count


def _write_categories(
    session: Session, categories: dict[str, dict[str, str | None]]
) -> int:
    count = 0
    for cid, fields in categories.items():
        session.add(
            BookCategory(
                id=cid,
                name_en=fields.get("name_en"),
                name_ar=fields.get("name_ar"),
                prefix=fields.get("prefix") or cid,
            )
        )
        count += 1
    session.flush()
    return count


def _write_books(session: Session, rows: Iterable[dict[str, Any]], known: set[str]) -> int:
    count = 0
    for r in rows:
        # Drop dangling employee_id references — v3 sometimes wrote "" or a
        # G-number that's been removed from the spreadsheet.
        if r.get("employee_id") and r["employee_id"] not in known:
            r = {**r, "employee_id": None}
        session.add(Book(**r))
        count += 1
    session.flush()
    return count


def _write_book_ref_sequence(session: Session, next_value: int) -> None:
    session.add(BookRefSequence(id=REF_SEQUENCE_ID, next_value=max(1, next_value)))
    session.flush()


def _write_leaves(session: Session, rows: Iterable[dict[str, Any]], known: set[str]) -> int:
    count = 0
    for r in rows:
        if r["employee_id"] not in known:
            continue
        session.add(Leave(**r))
        count += 1
    session.flush()
    return count


def _write_violations(
    session: Session, rows: Iterable[dict[str, Any]], known: set[str]
) -> int:
    count = 0
    for r in rows:
        if r["employee_id"] not in known:
            continue
        session.add(Violation(**r))
        count += 1
    session.flush()
    return count


def _write_managers(session: Session, rows: Iterable[dict[str, Any]]) -> int:
    count = 0
    for r in rows:
        # Drop id so SQLite auto-assigns one — preserves v3 ordering w/o conflict.
        r = {k: v for k, v in r.items() if k != "id" or v is not None}
        session.add(Manager(**r))
        count += 1
    session.flush()
    return count


def _write_settings(session: Session, mapping: dict[str, str]) -> int:
    count = 0
    for key, value in mapping.items():
        session.add(AppSetting(key=key, value=value))
        count += 1
    session.flush()
    return count


def _write_vault_files(session: Session, rows: Iterable[dict[str, Any]]) -> int:
    count = 0
    for r in rows:
        session.add(VaultFile(**r))
        count += 1
    session.flush()
    return count


# ----------------------------------------------------------------------------
# Orchestrator
# ----------------------------------------------------------------------------


def run_import(
    src_dir: Path,
    db_url: str,
    *,
    backup_root: Path,
    force: bool = False,
    dry: bool = False,
) -> ImportSummary:
    """Top-level entry point. Idempotent only via ``force`` — see module docstring."""
    summary = ImportSummary()

    if not src_dir.exists():
        raise FileNotFoundError(f"v3 source directory not found: {src_dir}")

    # 1. Backup before any reads.
    if not dry:
        summary.backup_dir = create_backup(src_dir, backup_root)

    # 2. Parse every source upfront so we can fail fast on a corrupt file.
    employees, skipped_emp = parse_employees(src_dir / "employees.xlsx")
    managers, skipped_mgr = parse_managers(src_dir / "managers.json")
    leaves, skipped_lv = parse_leaves(src_dir / "leave_history.json")
    violations, skipped_vio = parse_violations(src_dir / "violations.json")
    books_payload, skipped_bk = parse_books(src_dir / "books_database.json")
    settings, skipped_set = parse_settings(src_dir / "settings.json")
    summary.skipped.extend(
        skipped_emp + skipped_mgr + skipped_lv + skipped_vio + skipped_bk + skipped_set
    )

    known_employees: set[str] = {e["id"] for e in employees}
    vault_files, skipped_vf = walk_vault_files(
        src_dir / "employee_files", known_employees
    )
    summary.skipped.extend(skipped_vf)

    summary.employees = len(employees)
    summary.managers = len(managers)
    summary.leaves = sum(1 for r in leaves if r["employee_id"] in known_employees)
    summary.violations = sum(1 for r in violations if r["employee_id"] in known_employees)
    summary.book_categories = len(books_payload["categories"])
    summary.books = len(books_payload["books"])
    summary.settings = len(settings)
    summary.vault_files = len(vault_files)
    summary.next_ref_number = int(books_payload["next_ref_number"])

    if dry:
        return summary

    # 3. Open destination and ensure schema exists.
    engine = create_engine(db_url, future=True)
    is_memory = db_url.endswith(":memory:") or "mode=memory" in db_url
    attach_sqlite_pragmas(engine, wal=not is_memory)
    try:
        Base.metadata.create_all(engine)
        SessionFactory = sessionmaker(bind=engine, future=True, expire_on_commit=False)
        session = SessionFactory()
        try:
            if not _db_is_empty(session):
                if not force:
                    raise RuntimeError(
                        "destination DB already has parity rows; pass force=True to wipe"
                    )
                _truncate(session)

            _write_employees(session, employees)
            _write_categories(session, books_payload["categories"])
            _write_book_ref_sequence(session, summary.next_ref_number)
            _write_books(session, books_payload["books"], known_employees)
            _write_leaves(session, leaves, known_employees)
            _write_violations(session, violations, known_employees)
            _write_managers(session, managers)
            _write_settings(session, settings)
            _write_vault_files(session, vault_files)
            session.commit()
        finally:
            session.close()
    finally:
        _dispose(engine)

    return summary


def _dispose(engine: Engine) -> None:
    engine.dispose()


__all__ = [
    "ImportSummary",
    "create_backup",
    "parse_books",
    "parse_employees",
    "parse_leaves",
    "parse_managers",
    "parse_settings",
    "parse_violations",
    "run_import",
    "walk_vault_files",
]
