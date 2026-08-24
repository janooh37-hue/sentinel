from datetime import date

from app.db.models import Employee, TimesheetDesignation, TimesheetRosterAssignment
from app.services import timesheet_service as svc
from scripts import import_timesheet_history_2026 as importer


def _plan(designation_id: int) -> importer.Plan:
    return importer.Plan(
        new_leaves=[],
        new_absences=[],
        corrections=[],
        retypes=[],
        deletions=[],
        end_date_fixes=[],
        duplicate=None,
        duplicate_refs={},
        designations=[("G1001", date(2026, 8, 1), designation_id, "designation")],
        unmatched_designations={},
    )


def test_import_upserts_assignment_by_employee_and_effective_month(db_session):
    db_session.add(
        Employee(
            id="G1001",
            name_en="Name G1001",
            nationality="الإمارات",
            doj=date(2020, 1, 1),
        )
    )
    db_session.commit()
    svc.seed_designations(db_session)
    guard = db_session.query(TimesheetDesignation).filter_by(system_key="security_guard").one()
    supervisor = (
        db_session.query(TimesheetDesignation).filter_by(system_key="security_supervisor").one()
    )

    importer.apply_plan(db_session, _plan(guard.id))
    importer.apply_plan(db_session, _plan(supervisor.id))

    rows = db_session.query(TimesheetRosterAssignment).filter_by(employee_id="G1001").all()
    assert len(rows) == 1
    assert (rows[0].effective_from, rows[0].designation_id) == (date(2026, 8, 1), supervisor.id)


def _designation_book(path, employee_id, designation):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(importer.FIRST_DATA_ROW, importer.ID_COL, employee_id)
    sheet.cell(importer.FIRST_DATA_ROW, importer.DESIGNATION_COL, designation)
    workbook.save(path)


def test_import_reader_retains_each_workbook_month(monkeypatch, tmp_path):
    january = tmp_path / "january.xlsx"
    february = tmp_path / "february.xlsx"
    drivers = tmp_path / "drivers.xlsx"
    _designation_book(january, "G1001", "Security Guard")
    _designation_book(february, "G1001", "Security Supervisor")
    _designation_book(drivers, "G2000", "Driver")
    monkeypatch.setattr(importer, "SHARE", tmp_path)
    monkeypatch.setattr(importer, "MONTH_FILES", {1: january.name, 2: february.name})
    monkeypatch.setattr(importer, "DRIVERS_FILE", drivers.name)

    found = importer.read_designations()

    assert found == {
        "G1001": {
            date(2026, 1, 1): "Security Guard",
            date(2026, 2, 1): "Security Supervisor",
        },
        "G2000": {date(2026, 7, 1): "Driver"},
    }


def test_import_reader_skips_blank_cells_in_later_workbooks(monkeypatch, tmp_path):
    january = tmp_path / "january.xlsx"
    february = tmp_path / "february.xlsx"
    drivers = tmp_path / "drivers.xlsx"
    _designation_book(january, "G1001", "Security Guard")
    _designation_book(february, "G1001", None)
    _designation_book(drivers, "G2000", "Driver")
    monkeypatch.setattr(importer, "SHARE", tmp_path)
    monkeypatch.setattr(importer, "MONTH_FILES", {1: january.name, 2: february.name})
    monkeypatch.setattr(importer, "DRIVERS_FILE", drivers.name)

    found = importer.read_designations()

    assert found["G1001"] == {date(2026, 1, 1): "Security Guard"}


def test_duplicate_merge_preserves_noncolliding_assignment_history(db_session):
    db_session.add_all(
        [
            Employee(id="5704", name_en="Drop", nationality="الإمارات", doj=date(2020, 1, 1)),
            Employee(id="G5704", name_en="Keep", nationality="الإمارات", doj=date(2020, 1, 1)),
        ]
    )
    db_session.commit()
    svc.seed_designations(db_session)
    guard = db_session.query(TimesheetDesignation).filter_by(system_key="security_guard").one()
    supervisor = (
        db_session.query(TimesheetDesignation).filter_by(system_key="security_supervisor").one()
    )
    db_session.add_all(
        [
            TimesheetRosterAssignment(
                employee_id="5704",
                designation_id=guard.id,
                effective_from=date(2026, 2, 1),
            ),
            TimesheetRosterAssignment(
                employee_id="5704",
                designation_id=guard.id,
                effective_from=date(2026, 1, 1),
            ),
            TimesheetRosterAssignment(
                employee_id="G5704",
                designation_id=supervisor.id,
                effective_from=date(2026, 1, 1),
            ),
        ]
    )
    db_session.commit()
    plan = importer.Plan(
        new_leaves=[],
        new_absences=[],
        corrections=[],
        retypes=[],
        deletions=[],
        end_date_fixes=[],
        duplicate=("5704", "G5704", None),
        duplicate_refs={},
        designations=[],
        unmatched_designations={},
    )

    importer.apply_plan(db_session, plan)

    assert db_session.get(Employee, "5704") is None
    rows = (
        db_session.query(TimesheetRosterAssignment)
        .filter_by(employee_id="G5704")
        .order_by(TimesheetRosterAssignment.effective_from)
        .all()
    )
    assert [(row.effective_from, row.designation_id) for row in rows] == [
        (date(2026, 1, 1), supervisor.id),
        (date(2026, 2, 1), guard.id),
    ]
