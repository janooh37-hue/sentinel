# backend/tests/test_timesheet_template.py
"""The template is the format contract: if it drifts, every sheet drifts."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

TEMPLATE = Path(__file__).parents[1] / "templates" / "GSSG-HR_Monthly_Time_Sheet.xlsx"


@pytest.fixture(scope="module")
def workbook():
    return load_workbook(TEMPLATE)


def test_logo_survives(workbook):
    """The July attendance file on the share lost its logo to an outside tool."""
    assert workbook.sheetnames == ["Sheet1", "_parts"]
    assert len(workbook["Sheet1"]._images) == 1


def test_header_text_is_verbatim(workbook):
    sheet = workbook["Sheet1"]
    assert sheet["D1"].value == "Global Security Service Group- MONTHLY  TIME SHEET"
    assert sheet["Q2"].value == " Site Name :   JD 908"
    assert sheet["D3"].value == "Clent Code : P0331_JD_PRN_908EXT"
    assert sheet["E5"].value == "Desigantion"  # misspelled in the circulating sheets
    assert sheet["AO5"].value == "SL "  # trailing space drives the sick-leave COUNTIF


def test_day_headers_are_1_to_31(workbook):
    sheet = workbook["Sheet1"]
    assert [sheet.cell(5, 6 + i).value for i in range(31)] == list(range(1, 32))


def test_column_widths_match_the_reference(workbook):
    widths = workbook["Sheet1"].column_dimensions
    assert round(widths["C"].width, 2) == 69.43
    assert round(widths["E"].width, 2) == 39.71


def test_nothing_survives_below_the_header(workbook):
    """Clearing values would leave June's fills and borders on ~300 empty rows."""
    sheet = workbook["Sheet1"]
    assert sheet.max_row == 5
    assert not [r for r in sheet.merged_cells.ranges if r.min_row > 5]


def test_the_template_carries_no_conditional_formats_or_validations(workbook):
    """The renderer builds both from scratch over the real extent."""
    sheet = workbook["Sheet1"]
    assert list(sheet.conditional_formatting) == []
    assert sheet.data_validations.dataValidation == []


def test_parts_sheet_is_hidden_and_carries_the_19_row_footer(workbook):
    parts = workbook["_parts"]
    assert parts.sheet_state == "hidden"
    assert parts["A1"].font.name == "Arial"  # specimen data row
    assert "Legend:" in str(parts["A3"].value)
    assert str(parts["A4"].value).startswith("Prepard By")
    assert parts["A9"].value == "S.no"
    assert parts["A21"].value == "Total Days"


def test_the_red_block_has_a_legend_entry_and_a_footer_row(workbook):
    """A code the client cannot look up is worse than no code at all."""
    parts = workbook["_parts"]
    assert "X- Not billed" in str(parts["A3"].value)
    assert parts["C20"].value == "Not billed"
    assert parts["D20"].value == "X"
    # the new row borrows the OFF row's styling, so the block still reads as one table.
    # `border.top.style` alone is None on both and cannot fail; pin the style index and
    # a concrete border an unstyled cell would not have.
    assert parts["C20"]._style == parts["C19"]._style
    assert parts["C20"].border.bottom.style == "medium"


def test_the_red_block_row_is_code_height_not_summary_height(workbook):
    """At Total Days' height it reads as a second total rather than an eleventh code."""
    parts = workbook["_parts"]
    assert parts.row_dimensions[20].height == parts.row_dimensions[19].height
    assert parts.row_dimensions[20].height != parts.row_dimensions[21].height


def test_the_ten_code_rows_are_verbatim(workbook):
    """The paper's typos and trailing spaces are load-bearing: 'SL ' drives the COUNTIF."""
    parts = workbook["_parts"]
    assert [(parts.cell(r, 3).value, parts.cell(r, 4).value) for r in range(10, 20)] == [
        ("Sick Leave", "SL "),
        ("Annual Leave", "AL"),
        ("Abcent  ", "AB"),
        ("National Service", "TR"),
        ("New Gard", "NG"),
        ("Termination", "-"),
        ("Resignation", "R"),
        ("Suspention", "S "),
        ("P", "P"),
        ("OFF", "OFF"),
    ]


def test_no_june_filter_artifacts_survive(workbook):
    """June's autofilter is D5:E290 — a range over rows the strip deleted."""
    sheet = workbook["Sheet1"]
    assert sheet.auto_filter.ref is None
    # the four .wvu.FilterData names are sheet-local; the workbook collection is empty
    assert list(sheet.defined_names) == []
    assert list(workbook.defined_names) == []
