"""Routes, capability gates, and the freeze-on-download contract."""

from __future__ import annotations

import io
from dataclasses import fields
from datetime import date
from typing import get_args
from urllib.parse import quote

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from app.api.deps import get_current_user
from app.core import permissions, timesheet_xlsx
from app.core.constants import ARABIC_MONTHS
from app.db import session as session_mod
from app.db.models import (
    Absence,
    Base,
    Employee,
    Leave,
    TimesheetDesignation,
    TimesheetPeriod,
    TimesheetRosterAssignment,
    TimesheetSnapshotRow,
    TimesheetStartAck,
    TimesheetStatFiller,
    User,
    UserPermission,
)
from app.db.session import attach_sqlite_pragmas, get_db
from app.main import create_app
from app.schemas import timesheet as schemas
from app.services import perm_service
from app.services import timesheet_service as svc

XLSX_PREFIX = "application/vnd.openxmlformats-officedocument.spreadsheetml"


@pytest.fixture()
def api_db(monkeypatch, tmp_path) -> Session:
    db_file = tmp_path / "timesheet.db"
    eng = create_engine(
        f"sqlite:///{db_file}", future=True, connect_args={"check_same_thread": False}
    )
    attach_sqlite_pragmas(eng, wal=False)
    Base.metadata.create_all(eng)
    TestSession = sessionmaker(bind=eng, autoflush=False, expire_on_commit=False, future=True)
    monkeypatch.setattr(session_mod, "engine", eng)
    monkeypatch.setattr(session_mod, "SessionLocal", TestSession)
    db = TestSession()
    perm_service.seed_role_defaults(db)
    svc.seed_designations(db)  # metadata.create_all skips the migration seed
    try:
        yield db
    finally:
        db.close()


@pytest.fixture()
def db_session(api_db) -> Session:
    """Shadow the conftest fixture so the client and the seeds share one DB."""
    return api_db


def _client_for(api_db: Session, role: str, email: str) -> TestClient:
    user = User(email=email, password_hash="x", role=role, status="active")
    api_db.add(user)
    api_db.commit()
    api_db.refresh(user)
    app = create_app()
    app.dependency_overrides[get_db] = lambda: api_db
    app.dependency_overrides[get_current_user] = lambda: user
    client = TestClient(app, raise_server_exceptions=True)
    client.user_id = user.id  # type: ignore[attr-defined]
    return client


@pytest.fixture()
def client(api_db) -> TestClient:
    """A manager: `timesheet.view` (inherited) + `timesheet.edit`."""
    return _client_for(api_db, "manager", "mgr@x.ae")


@pytest.fixture()
def viewer_client(api_db) -> TestClient:
    """An operator: `timesheet.view` only — the read-only half of the page."""
    return _client_for(api_db, "operator", "ops@x.ae")


def _guard(db: Session) -> None:
    designation = db.query(TimesheetDesignation).filter_by(name_en="Security Guard").one()
    db.add(
        Employee(
            id="G1001",
            name_en="TEST GUARD",
            nationality="الإمارات",
            doj=date(2020, 1, 1),
        )
    )
    db.add(
        TimesheetRosterAssignment(
            employee_id="G1001",
            designation_id=designation.id,
            effective_from=date(2026, 1, 1),
        )
    )
    db.commit()


def _driver(db: Session) -> None:
    designation = db.query(TimesheetDesignation).filter_by(name_en="Driver").one()
    db.add(
        Employee(
            id="G2000",
            name_en="TEST DRIVER",
            nationality="الإمارات",
            doj=date(2020, 1, 1),
        )
    )
    db.add(
        TimesheetRosterAssignment(
            employee_id="G2000",
            designation_id=designation.id,
            effective_from=date(2026, 1, 1),
        )
    )
    db.commit()


def _add_assignment(
    db: Session,
    employee_id: str,
    designation_id: int | None,
    effective_from: date = date(2026, 1, 1),
) -> None:
    db.add(
        TimesheetRosterAssignment(
            employee_id=employee_id,
            designation_id=designation_id,
            effective_from=effective_from,
        )
    )


# --------------------------------------------------------------------------- #
# the grid
# --------------------------------------------------------------------------- #


def test_get_returns_the_grid(client, db_session):
    _guard(db_session)
    body = client.get("/api/v1/timesheet/2026/7").json()
    assert body["days_in_month"] == 31
    assert body["post_count"] == 249
    assert body["rows"][0]["employee_id"] == "G1001"
    assert body["rows"][0]["codes"][0] == "P"


def test_the_grid_response_loses_no_field(client, db_session):
    """Every ``MonthGrid`` and ``GridRow`` field reaches the page.

    Asserted against the dataclasses, so a field added to the service and
    forgotten in the schema fails here instead of silently disabling a feature.
    """

    _guard(db_session)
    client.put(
        "/api/v1/timesheet/2026/7/cell",
        json={"employee_id": "G1001", "day": 9, "code": "AB", "note": "no show"},
    )
    body = client.get("/api/v1/timesheet/2026/7").json()

    grid_fields = {f.name for f in fields(svc.MonthGrid)}
    row_fields = {f.name for f in fields(svc.GridRow)}
    assert len(grid_fields) == 11 and len(row_fields) == 16
    assert set(body) == grid_fields
    row = body["rows"][0]
    assert set(row) == row_fields
    assert row["designation_id"] == db_session.query(TimesheetRosterAssignment).one().designation_id
    assert row["stat_filler"] is None
    assert row["joined_day"] is None
    assert row["left_day"] is None
    assert row["start_confirmed"] is False


def test_notes_serialise_with_string_keys(client, db_session):
    """``dict[int, str]`` in Python; the page indexes it as ``notes[String(day)]``."""

    _guard(db_session)
    client.put(
        "/api/v1/timesheet/2026/7/cell",
        json={"employee_id": "G1001", "day": 9, "code": "AB", "note": "no show"},
    )
    notes = client.get("/api/v1/timesheet/2026/7").json()["rows"][0]["notes"]
    assert notes == {"9": "no show"}
    assert [type(key) for key in notes] == [str]


def test_a_warning_may_name_someone_with_no_row(client, db_session):
    """``warnings`` and ``removed`` are top-level, never nested in a row."""

    designation = db_session.query(TimesheetDesignation).filter_by(name_en="Security Guard").one()
    db_session.add(
        Employee(
            id="G1002",
            name_en="GONE GUARD",
            nationality="الإمارات",
            doj=date(2020, 1, 1),
            end_date=date(2026, 6, 20),
            status="Active",
        )
    )
    _add_assignment(db_session, "G1002", designation.id)
    db_session.commit()

    body = client.get("/api/v1/timesheet/2026/7").json()
    assert [row["employee_id"] for row in body["rows"]] == []
    warning = next(item for item in body["warnings"] if item["kind"] == "departed_but_active")
    assert warning["employee_id"] == "G1002"
    assert set(warning) == {"employee_id", "kind", "detail"}
    removed = body["removed"][0]
    assert set(removed) == {f.name for f in fields(svc.Removed)}
    assert (removed["employee_id"], removed["last_day"], removed["month"], removed["year"]) == (
        "G1002",
        20,
        6,
        2026,
    )


def test_an_impossible_month_is_refused_not_a_500(client):
    assert client.get("/api/v1/timesheet/2026/13").status_code == 422
    assert client.get("/api/v1/timesheet/2026/0").status_code == 422


def test_the_drivers_sheet_is_its_own_grid(client, db_session):
    _guard(db_session)
    _driver(db_session)
    assert [r["employee_id"] for r in client.get("/api/v1/timesheet/2026/7").json()["rows"]] == [
        "G1001"
    ]
    body = client.get("/api/v1/timesheet/2026/7?sheet=drivers").json()
    assert [r["employee_id"] for r in body["rows"]] == ["G2000"]
    assert body["sheet"] == "drivers"


# --------------------------------------------------------------------------- #
# editing
# --------------------------------------------------------------------------- #


def test_put_cell_marks_absence(client, db_session):
    _guard(db_session)
    response = client.put(
        "/api/v1/timesheet/2026/7/cell",
        json={"employee_id": "G1001", "day": 9, "code": "AB", "note": "no show"},
    )
    assert response.status_code == 200
    assert client.get("/api/v1/timesheet/2026/7").json()["rows"][0]["codes"][8] == "AB"


def test_a_cell_edit_records_who_made_it(client, db_session):
    _guard(db_session)
    client.put(
        "/api/v1/timesheet/2026/7/cell", json={"employee_id": "G1001", "day": 9, "code": "AB"}
    )
    assert db_session.query(Absence).one().created_by == client.user_id


def test_a_bad_cell_code_answers_with_the_service_error(client, db_session):
    _guard(db_session)
    response = client.put(
        "/api/v1/timesheet/2026/7/cell", json={"employee_id": "G1001", "day": 9, "code": "ZZ"}
    )
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "TIMESHEET_BAD_CODE"


def test_a_cell_edit_for_an_unknown_employee_is_a_404(client, db_session):
    _guard(db_session)
    response = client.put(
        "/api/v1/timesheet/2026/7/cell", json={"employee_id": "G9998", "day": 9, "code": "AB"}
    )
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "EMPLOYEE_NOT_FOUND"


def test_patch_sets_the_post_count_and_a_driver_filler(client, db_session):
    _driver(db_session)
    response = client.patch(
        "/api/v1/timesheet/2026/7?sheet=drivers",
        json={"post_count": 0, "fillers": [{"employee_id": "G2000", "code": "SL "}]},
    )
    assert response.status_code == 200
    row = response.json()["rows"][0]
    assert row["stat_block"] == 2
    assert row["stat_codes"][0] == "SL "


def test_a_patch_whose_filler_fails_changes_nothing(client, db_session):
    """One unit of work: both writers commit on their own by default."""

    _guard(db_session)
    response = client.patch(
        "/api/v1/timesheet/2026/7",
        json={
            "post_count": 240,
            "fillers": [
                {"employee_id": "G1001", "code": "SL "},
                {"employee_id": "G9998", "code": "AL"},  # no such employee
            ],
        },
    )
    assert response.status_code == 404
    db_session.rollback()  # the route's failed transaction, seen from the test session
    assert db_session.query(TimesheetPeriod).count() == 0
    assert db_session.query(TimesheetStatFiller).count() == 0
    assert client.get("/api/v1/timesheet/2026/7").json()["post_count"] == 249


def test_a_closed_month_rejects_a_cell_edit(client, db_session):
    _guard(db_session)
    client.get("/api/v1/timesheet/2026/7/export")
    response = client.put(
        "/api/v1/timesheet/2026/7/cell", json={"employee_id": "G1001", "day": 9, "code": "AB"}
    )
    assert response.status_code == 409


def test_reopen_unlocks_it(client, db_session):
    _guard(db_session)
    client.get("/api/v1/timesheet/2026/7/export")
    assert client.post("/api/v1/timesheet/2026/7/reopen").status_code == 200
    assert (
        client.put(
            "/api/v1/timesheet/2026/7/cell", json={"employee_id": "G1001", "day": 9, "code": "AB"}
        ).status_code
        == 200
    )


def test_close_and_reopen_record_the_actor(client, db_session):
    _guard(db_session)
    body = client.post("/api/v1/timesheet/2026/7/close").json()
    assert body["closed_at"] is not None
    assert body["closed_by"] == "mgr@x.ae"
    period = db_session.query(TimesheetPeriod).one()
    assert period.closed_by == client.user_id
    client.post("/api/v1/timesheet/2026/7/reopen")
    db_session.refresh(period)
    assert period.closed_at is None
    assert period.reopened_by == client.user_id


# --------------------------------------------------------------------------- #
# the starting-point acknowledgement
# --------------------------------------------------------------------------- #


def test_start_ack_is_204_idempotent_and_records_the_actor(client, db_session):
    designation = db_session.query(TimesheetDesignation).filter_by(name_en="Security Guard").one()
    db_session.add(
        Employee(
            id="G7176",
            name_en="NEW GUARD",
            nationality="الإمارات",
            doj=date(2026, 7, 14),
        )
    )
    _add_assignment(db_session, "G7176", designation.id)
    db_session.commit()

    first = client.post("/api/v1/timesheet/2026/7/start-ack", json={"employee_id": "G7176"})
    assert first.status_code == 204
    assert first.content == b""
    assert (
        client.post("/api/v1/timesheet/2026/7/start-ack", json={"employee_id": "G7176"}).status_code
        == 204
    )
    ack = db_session.query(TimesheetStartAck).one()
    assert ack.acked_by == client.user_id
    row = client.get("/api/v1/timesheet/2026/7").json()["rows"][0]
    assert row["start_confirmed"] is True
    assert row["joined_day"] == 14


def test_start_ack_is_allowed_on_a_closed_month(client, db_session):
    """Refusing it after the close would strand the flag forever."""

    _guard(db_session)
    client.get("/api/v1/timesheet/2026/7/export")
    assert client.get("/api/v1/timesheet/2026/7").json()["closed_at"] is not None
    assert (
        client.post("/api/v1/timesheet/2026/7/start-ack", json={"employee_id": "G1001"}).status_code
        == 204
    )
    assert db_session.query(TimesheetStartAck).count() == 1


def test_start_ack_for_an_unknown_employee_is_a_404(client, db_session):
    _guard(db_session)
    response = client.post("/api/v1/timesheet/2026/7/start-ack", json={"employee_id": "G9999"})
    assert response.status_code == 404
    # The service's code, not Starlette's HTTP_404: a bare status assertion here
    # passes against a route that does not exist at all.
    assert response.json()["error"]["code"] == "EMPLOYEE_NOT_FOUND"
    assert db_session.query(TimesheetStartAck).count() == 0


def test_start_ack_for_someone_off_the_roster_is_a_404(client, db_session):
    """Stored, it would render ``start_confirmed`` over a start nobody accepted.

    ``acknowledge_start`` only checks that the employee exists and
    ``timesheet_start_acks`` has no foreign key, so the roster clause has to be
    enforced on the way in.
    """

    designation = db_session.query(TimesheetDesignation).filter_by(name_en="Security Guard").one()
    db_session.add(
        Employee(
            id="G3000",
            name_en="FUTURE GUARD",
            nationality="الإمارات",
            doj=date(2026, 9, 1),  # exists, but not on July's roster
        )
    )
    _add_assignment(db_session, "G3000", designation.id)
    db_session.commit()

    response = client.post("/api/v1/timesheet/2026/7/start-ack", json={"employee_id": "G3000"})
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "EMPLOYEE_NOT_ON_SHEET"
    assert db_session.query(TimesheetStartAck).count() == 0
    # …and he is accepted for the month he does join.
    assert (
        client.post("/api/v1/timesheet/2026/9/start-ack", json={"employee_id": "G3000"}).status_code
        == 204
    )


# --------------------------------------------------------------------------- #
# the workbooks
# --------------------------------------------------------------------------- #


def test_export_returns_an_xlsx_with_an_rfc5987_arabic_filename(client, db_session):
    _guard(db_session)
    response = client.get("/api/v1/timesheet/2026/7/export?variant=attendance")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(XLSX_PREFIX)
    disposition = response.headers["content-disposition"]
    assert disposition.startswith("attachment;")
    assert "filename*=UTF-8''" in disposition  # a bare filename= raises on latin-1
    assert quote("كشف حضور") in disposition
    assert client.get("/api/v1/timesheet/2026/7").json()["closed_at"] is not None


def test_the_statistics_variant_is_a_second_workbook(client, db_session):
    _guard(db_session)
    response = client.get("/api/v1/timesheet/2026/7/export?variant=statistics")
    assert response.status_code == 200
    assert quote("الاحصائية شهر") in response.headers["content-disposition"]


def test_an_unknown_variant_is_refused(client, db_session):
    _guard(db_session)
    assert client.get("/api/v1/timesheet/2026/7/export?variant=nonsense").status_code == 422


def test_export_blocks_when_an_employee_has_no_designation(client, db_session):
    db_session.add(
        Employee(id="G9999", name_en="Nobody", nationality="الإمارات", doj=date(2020, 1, 1))
    )
    db_session.commit()
    response = client.get("/api/v1/timesheet/2026/7/export")
    assert response.status_code == 422
    assert "no_designation" in response.text
    assert db_session.query(TimesheetPeriod).count() == 0  # a refused download freezes nothing


def test_export_preflights_both_sheets_before_sealing(client, db_session):
    """A clean requested sheet must not seal a different blocked sheet."""

    _guard(db_session)
    designation = db_session.query(TimesheetDesignation).filter_by(name_en="Driver").one()
    db_session.add(
        Employee(
            id="G2000",
            name_en="UNKNOWN NATIONALITY DRIVER",
            nationality="India",
            doj=date(2020, 1, 1),
        )
    )
    _add_assignment(db_session, "G2000", designation.id)
    db_session.commit()

    response = client.get("/api/v1/timesheet/2026/7/export")
    assert response.status_code == 422
    assert response.json()["error"]["message"] == (
        "Fix the blocking issues on the following sheet(s) before downloading: drivers."
    )
    blocking = response.json()["error"]["details"]["blocking"]
    assert {issue["sheet"] for issue in blocking} == {"drivers"}
    assert all(issue["kind"] == "no_nationality" for issue in blocking)
    assert db_session.query(TimesheetPeriod).count() == 0
    assert db_session.query(TimesheetSnapshotRow).count() == 0

    close = client.post("/api/v1/timesheet/2026/7/close")
    assert close.status_code == 422
    assert {issue["sheet"] for issue in close.json()["error"]["details"]["blocking"]} == {"drivers"}
    assert db_session.query(TimesheetPeriod).count() == 0
    assert db_session.query(TimesheetSnapshotRow).count() == 0


def test_single_employee_export(client, db_session):
    _guard(db_session)
    response = client.get("/api/v1/timesheet/employee/G1001/2026/7/export")
    assert response.status_code == 200
    assert "filename*=UTF-8''" in response.headers["content-disposition"]


def test_the_employee_export_freezes_nothing(client, db_session):
    _guard(db_session)
    client.get("/api/v1/timesheet/employee/G1001/2026/7/export")
    assert client.get("/api/v1/timesheet/2026/7").json()["closed_at"] is None


def test_the_employee_export_finds_a_driver_on_his_own_sheet(client, db_session):
    _driver(db_session)
    response = client.get("/api/v1/timesheet/employee/G2000/2026/7/export")
    assert response.status_code == 200
    assert load_workbook(io.BytesIO(response.content)).worksheets[0]["B6"].value == "G2000"


def test_employee_export_uses_the_requested_month_assignment(client, db_session):
    _guard(db_session)
    july = client.get("/api/v1/timesheet/employee/G1001/2026/7/export")
    assert load_workbook(io.BytesIO(july.content)).worksheets[0]["B6"].value == "G1001"

    driver = db_session.query(TimesheetDesignation).filter_by(name_en="Driver").one()
    _add_assignment(db_session, "G1001", driver.id, date(2026, 8, 1))
    db_session.commit()
    august = client.get("/api/v1/timesheet/employee/G1001/2026/8/export")
    assert load_workbook(io.BytesIO(august.content)).worksheets[0]["B6"].value == "G1001"


def test_two_month_export_resolves_each_month_sheet_after_assignment_change(client, db_session):
    _guard(db_session)
    driver = db_session.query(TimesheetDesignation).filter_by(name_en="Driver").one()
    _add_assignment(db_session, "G1001", driver.id, date(2026, 8, 1))
    db_session.commit()

    response = client.get("/api/v1/timesheet/employee/G1001/2026/8/export?months=2")
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["JUL", "AUG"]
    assert workbook["JUL"]["B6"].value == "G1001"
    assert workbook["AUG"]["B6"].value == "G1001"


def test_two_months_are_one_workbook_earlier_sheet_first(client, db_session):
    _guard(db_session)
    response = client.get("/api/v1/timesheet/employee/G1001/2026/7/export?months=2")
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["JUN", "JUL"]
    assert workbook["JUN"]["B6"].value == "G1001"
    assert workbook["JUL"]["B6"].value == "G1001"
    # The name carries the LATER month, which is the month of departure.
    disposition = response.headers["content-disposition"]
    assert quote(f"كشف حضور TEST GUARD {ARABIC_MONTHS[6]}") in disposition
    assert quote(ARABIC_MONTHS[5]) not in disposition


def test_two_months_step_back_over_the_december_boundary(client, db_session):
    _guard(db_session)
    response = client.get("/api/v1/timesheet/employee/G1001/2027/1/export?months=2")
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["DEC", "JAN"]
    assert workbook["DEC"]["D4"].value == "For the Month of :DEC-2026"
    assert workbook["JAN"]["D4"].value == "For the Month of :JAN-2027"


def test_the_handover_span_of_someone_who_left_last_month(client, db_session):
    """The case ``months=2`` exists for, asked from the month Task 10 is on.

    He finished in June, so July has no row for him. ``render_single_span``
    tolerates that by design, and the file takes the earlier month's name — an
    eager ``filename_for_single`` on the later grid 404s the whole download.
    """

    designation = db_session.query(TimesheetDesignation).filter_by(name_en="Security Guard").one()
    db_session.add(
        Employee(
            id="G4000",
            name_en="LEFT GUARD",
            nationality="الإمارات",
            doj=date(2020, 1, 1),
            end_date=date(2026, 6, 20),
        )
    )
    _add_assignment(db_session, "G4000", designation.id)
    db_session.commit()

    response = client.get("/api/v1/timesheet/employee/G4000/2026/7/export?months=2")
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["JUN", "JUL"]
    assert workbook["JUN"]["B6"].value == "G4000"
    assert workbook["JUL"]["B6"].value is None
    assert (
        quote(f"كشف حضور LEFT GUARD {ARABIC_MONTHS[5]}")
        in (response.headers["content-disposition"])
    )


def test_a_span_that_misses_both_months_still_404s(client, db_session):
    designation = db_session.query(TimesheetDesignation).filter_by(name_en="Security Guard").one()
    db_session.add(
        Employee(
            id="G5000",
            name_en="LATER GUARD",
            nationality="الإمارات",
            doj=date(2026, 10, 1),
        )
    )
    _add_assignment(db_session, "G5000", designation.id)
    db_session.commit()

    response = client.get("/api/v1/timesheet/employee/G5000/2026/7/export?months=2")
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "EMPLOYEE_NOT_ON_SHEET"


@pytest.mark.parametrize("months", [0, 3, 12])
def test_months_outside_one_and_two_is_refused(client, db_session, months):
    _guard(db_session)
    response = client.get(f"/api/v1/timesheet/employee/G1001/2026/7/export?months={months}")
    assert response.status_code == 422


def test_the_employee_export_404s_for_someone_off_the_sheet(client, db_session):
    _guard(db_session)
    unknown = client.get("/api/v1/timesheet/employee/G9999/2026/7/export")
    assert unknown.status_code == 404
    # The service's code, not Starlette's HTTP_404, which a missing route returns.
    assert unknown.json()["error"]["code"] == "EMPLOYEE_NOT_FOUND"

    designation = db_session.query(TimesheetDesignation).filter_by(name_en="Driver").one()
    db_session.add(
        Employee(
            id="G3000",
            name_en="LATER GUARD",
            nationality="الإمارات",
            doj=date(2026, 9, 1),  # employed, but not in July
        )
    )
    _add_assignment(db_session, "G3000", designation.id)
    db_session.commit()
    off_roster = client.get("/api/v1/timesheet/employee/G3000/2026/7/export")
    assert off_roster.status_code == 404
    assert off_roster.json()["error"]["code"] == "EMPLOYEE_NOT_ON_SHEET"


def test_the_static_query_types_match_the_runtime_constants():
    """``Sheet``/``Variant`` cannot be spelled from a tuple, so pin them here."""

    assert set(get_args(schemas.Sheet)) == set(svc.SHEETS)
    assert set(get_args(schemas.Variant)) == timesheet_xlsx.VARIANTS


# --------------------------------------------------------------------------- #
# the designation catalog
# --------------------------------------------------------------------------- #


def test_designations_list_and_reorder(client):
    body = client.get("/api/v1/timesheet/designations").json()
    assert len(body) == 16
    assert [d["rank_order"] for d in body] == list(range(1, 17))
    assert body[0]["name_en"] == "Prisons Director"
    assert body[-1]["sheet"] == "drivers"
    ids = [d["id"] for d in body]
    assert (
        client.put(
            "/api/v1/timesheet/designations/order", json={"ids": [ids[1], ids[0], *ids[2:]]}
        ).status_code
        == 200
    )
    assert client.get("/api/v1/timesheet/designations").json()[0]["name_en"] == "Ass. Director"
    assert (
        client.put("/api/v1/timesheet/designations/order", json={"ids": ids[:5]}).status_code == 422
    )


def test_designation_create_normalizes_names_and_rename_preserves_catalog_fields(
    client, db_session
):
    created = client.post(
        "/api/v1/timesheet/designations",
        json={"name_en": " Relief Supervisor ", "name_ar": " مشرف بديل ", "sheet": "main"},
    )
    assert created.status_code == 200
    body = created.json()
    assert (body["name_en"], body["name_ar"], body["system_key"]) == (
        "Relief Supervisor",
        "مشرف بديل",
        None,
    )
    assert body["rank_order"] == 17

    renamed = client.patch(
        f"/api/v1/timesheet/designations/{body['id']}",
        json={"name_en": " Relief Duty Supervisor ", "name_ar": " مشرف مناوب بديل "},
    )
    assert renamed.status_code == 200
    assert renamed.json()["name_en"] == "Relief Duty Supervisor"
    assert renamed.json()["name_ar"] == "مشرف مناوب بديل"
    assert renamed.json()["rank_order"] == 17
    assert renamed.json()["sheet"] == "main"
    assert renamed.json()["system_key"] is None

    duplicate = client.post(
        "/api/v1/timesheet/designations",
        json={"name_en": " security guard ", "name_ar": "حارس آخر", "sheet": "main"},
    )
    assert duplicate.status_code == 422
    assert duplicate.json()["error"]["code"] == "DESIGNATION_NAME_DUPLICATE"


def test_roster_put_upserts_effective_assignment_and_allows_explicit_null(client, db_session):
    _guard(db_session)
    designation = db_session.query(TimesheetDesignation).filter_by(name_en="Driver").one()
    response = client.put(
        "/api/v1/timesheet/2026/8/roster",
        json={"assignments": [{"employee_id": "G1001", "designation_id": designation.id}]},
    )
    assert response.status_code == 204
    row = (
        db_session.query(TimesheetRosterAssignment)
        .filter_by(employee_id="G1001", effective_from=date(2026, 8, 1))
        .one()
    )
    assert (row.designation_id, row.assigned_by) == (designation.id, client.user_id)

    cleared = client.put(
        "/api/v1/timesheet/2026/8/roster",
        json={"assignments": [{"employee_id": "G1001", "designation_id": None}]},
    )
    assert cleared.status_code == 204
    db_session.refresh(row)
    assert row.designation_id is None


def test_roster_duplicate_employee_is_atomic(client, db_session):
    _guard(db_session)
    before = [
        (row.employee_id, row.effective_from, row.designation_id)
        for row in db_session.query(TimesheetRosterAssignment).all()
    ]
    response = client.put(
        "/api/v1/timesheet/2026/8/roster",
        json={
            "assignments": [
                {"employee_id": "G1001", "designation_id": 1},
                {"employee_id": "G1001", "designation_id": None},
            ]
        },
    )
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "ROSTER_DUPLICATE_EMPLOYEE"
    db_session.rollback()
    assert [
        (row.employee_id, row.effective_from, row.designation_id)
        for row in db_session.query(TimesheetRosterAssignment).all()
    ] == before


def test_roster_unknown_employee_is_atomic(client, db_session):
    _guard(db_session)
    before = db_session.query(TimesheetRosterAssignment).count()
    designation = db_session.query(TimesheetDesignation).filter_by(name_en="Driver").one()
    response = client.put(
        "/api/v1/timesheet/2026/8/roster",
        json={
            "assignments": [
                {"employee_id": "G1001", "designation_id": designation.id},
                {"employee_id": "G9999", "designation_id": designation.id},
            ]
        },
    )
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "EMPLOYEE_NOT_FOUND"
    db_session.rollback()
    assert db_session.query(TimesheetRosterAssignment).count() == before


def test_roster_inactive_designation_is_atomic(client, db_session):
    _guard(db_session)
    designation = db_session.query(TimesheetDesignation).filter_by(name_en="Driver").one()
    designation.active = False
    db_session.commit()
    before = db_session.query(TimesheetRosterAssignment).count()
    response = client.put(
        "/api/v1/timesheet/2026/8/roster",
        json={"assignments": [{"employee_id": "G1001", "designation_id": designation.id}]},
    )
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "DESIGNATION_INACTIVE"
    db_session.rollback()
    assert db_session.query(TimesheetRosterAssignment).count() == before


def test_roster_closed_month_is_atomic(client, db_session):
    _guard(db_session)
    client.get("/api/v1/timesheet/2026/8/export")
    designation = db_session.query(TimesheetDesignation).filter_by(name_en="Driver").one()
    before = db_session.query(TimesheetRosterAssignment).count()
    response = client.put(
        "/api/v1/timesheet/2026/8/roster",
        json={"assignments": [{"employee_id": "G1001", "designation_id": designation.id}]},
    )
    assert response.status_code == 409
    assert response.json()["error"]["code"] == "TIMESHEET_CLOSED"
    db_session.rollback()
    assert db_session.query(TimesheetRosterAssignment).count() == before


def test_the_static_designation_routes_are_declared_before_the_month_route():
    """Static catalog and roster routes must not be shadowed by month paths."""

    paths = [getattr(route, "path", "") for route in create_app().routes]
    month = paths.index("/api/v1/timesheet/{year}/{month}")
    assert paths.index("/api/v1/timesheet/designations") < month
    assert paths.index("/api/v1/timesheet/designations/order") < month
    assert paths.index("/api/v1/timesheet/designations/{designation_id}") < month
    assert paths.index("/api/v1/timesheet/{year}/{month}/roster") < month


# --------------------------------------------------------------------------- #
# the capability gates
# --------------------------------------------------------------------------- #


def test_the_presets_split_reading_from_correcting():
    assert "timesheet.view" in permissions._OPERATOR_CAPS
    assert "timesheet.edit" not in permissions._OPERATOR_CAPS
    assert "timesheet.edit" in permissions._MANAGER_CAPS
    assert "timesheet.view" in permissions._MANAGER_CAPS  # inherited, never listed twice
    assert {"timesheet.view", "timesheet.edit"} <= permissions.ALL_CAPABILITIES
    assert {"timesheet.view", "timesheet.edit"} <= set(permissions.CAPABILITY_IDS)


@pytest.mark.parametrize(
    ("method", "path", "body"),
    [
        (
            "POST",
            "/api/v1/timesheet/designations",
            {"name_en": "New", "name_ar": "جديد", "sheet": "main"},
        ),
        ("PATCH", "/api/v1/timesheet/designations/1", {"name_en": "Renamed", "name_ar": "معدل"}),
        (
            "PUT",
            "/api/v1/timesheet/2026/7/roster",
            {"assignments": [{"employee_id": "G1001", "designation_id": 1}]},
        ),
        ("PUT", "/api/v1/timesheet/designations/order", {"ids": [1]}),
        ("PUT", "/api/v1/timesheet/2026/7/cell", {"employee_id": "G1001", "day": 9, "code": "AB"}),
        ("PATCH", "/api/v1/timesheet/2026/7", {"post_count": 100}),
        ("POST", "/api/v1/timesheet/2026/7/close", None),
        ("POST", "/api/v1/timesheet/2026/7/reopen", None),
        ("POST", "/api/v1/timesheet/2026/7/start-ack", {"employee_id": "G1001"}),
        ("GET", "/api/v1/timesheet/2026/7/export", None),
    ],
)
def test_view_only_is_refused_every_edit_route(viewer_client, db_session, method, path, body):
    """The month export included: downloading it freezes the month."""

    _guard(db_session)
    response = viewer_client.request(method, path, json=body)
    assert response.status_code == 403
    assert response.json()["error"]["details"]["capability"] == "timesheet.edit"
    assert db_session.query(TimesheetPeriod).count() == 0


def test_view_only_reads_the_grid_and_takes_one_employee_home(viewer_client, db_session):
    _guard(db_session)
    assert viewer_client.get("/api/v1/timesheet/2026/7").status_code == 200
    assert viewer_client.get("/api/v1/timesheet/designations").status_code == 200
    assert viewer_client.get("/api/v1/timesheet/employee/G1001/2026/7/export").status_code == 200
    assert viewer_client.get("/api/v1/timesheet/2026/7").json()["closed_at"] is None


@pytest.mark.parametrize(
    "path",
    [
        "/api/v1/timesheet/2026/7",
        "/api/v1/timesheet/designations",
        "/api/v1/timesheet/employee/G1001/2026/7/export",
    ],
)
def test_denying_view_closes_all_three_read_routes(viewer_client, db_session, path):
    """A refusal, not an admission: a 200 for an operator also passes with the
    ``require_capability("timesheet.view")`` dependency deleted.

    ``require_capability`` keeps the capability in a closure cell with nothing to
    introspect, so the gate is proven by taking the capability away —
    ``perm_service`` resolves role defaults plus grants minus denies.
    """

    _guard(db_session)
    db_session.add(
        UserPermission(user_id=viewer_client.user_id, capability="timesheet.view", effect="deny")
    )
    db_session.commit()
    response = viewer_client.get(path)
    assert response.status_code == 403
    assert response.json()["error"]["details"]["capability"] == "timesheet.view"


# --------------------------------------------------------------------------- #
# the absence supersede hook
# --------------------------------------------------------------------------- #


@pytest.fixture()
def generation_env(db_session, tmp_path, monkeypatch):
    """The two-line isolation every generation test uses.

    Unstubbed, ``generate_document`` drives Word COM through
    ``_pdf_executor.convert_docx_to_pdf`` (a 120 s process-pool wait); see
    ``backend/tests/test_document_generation_included_papers.py:33-63``.
    """

    from app.config import Settings
    from app.db.models import BookCategory
    from app.services import document_service

    settings = Settings(data_dir=tmp_path / "data")
    monkeypatch.setattr(document_service, "get_settings", lambda: settings)
    monkeypatch.setattr(document_service, "convert_docx_to_pdf", lambda p: None)
    db_session.add(BookCategory(id="HR", prefix="HR"))  # FK on Book.category_id
    db_session.commit()
    return document_service


def _sick_certificate(document_service, db, day=9):
    return document_service.generate_document(
        db,  # the only positional parameter
        employee_id="G1001",
        template_id="Leave Application Form",
        fields={
            "leave_type": "Sick Leave",
            "start_date": f"2026-07-{day:02d}",
            "end_date": f"2026-07-{day:02d}",
            "total_days": 1,
        },
        commit=True,
    )


def test_generating_a_sick_leave_clears_the_absence(client, db_session, generation_env):
    """Drives the real creation path — a hand-inserted Leave would not exercise the hook."""

    _guard(db_session)
    client.put(
        "/api/v1/timesheet/2026/7/cell", json={"employee_id": "G1001", "day": 9, "code": "AB"}
    )
    assert db_session.query(Absence).count() == 1
    result = _sick_certificate(generation_env, db_session)
    assert result.leave_id is not None
    assert db_session.query(Absence).count() == 0
    assert client.get("/api/v1/timesheet/2026/7").json()["rows"][0]["codes"][8] == "SL "


def test_regenerating_the_same_certificate_supersedes_it_again(client, db_session, generation_env):
    """The dedup branch reuses the row and never calls ``db.add`` — it must still fire."""

    _guard(db_session)
    first = _sick_certificate(generation_env, db_session)
    client.put(
        "/api/v1/timesheet/2026/7/cell", json={"employee_id": "G1001", "day": 9, "code": "AB"}
    )
    assert db_session.query(Absence).count() == 1

    again = _sick_certificate(generation_env, db_session)
    assert again.leave_id == first.leave_id  # proves the dedup path ran
    assert db_session.query(Absence).count() == 0


def test_a_leave_that_is_no_day_code_leaves_the_absence_alone(client, db_session, generation_env):
    """The gate is ``leave_code``, so a Passport Release supersedes nothing."""

    _guard(db_session)
    client.put(
        "/api/v1/timesheet/2026/7/cell", json={"employee_id": "G1001", "day": 9, "code": "AB"}
    )
    generation_env.generate_document(
        db_session,
        employee_id="G1001",
        template_id="Passport Release Form",
        fields={"request_date": "2026-07-09", "return_date": "2026-07-09"},
        commit=True,
    )
    assert db_session.query(Absence).count() == 1


def test_a_failure_after_step_12_leaves_the_absences_intact(
    client, db_session, generation_env, monkeypatch
):
    """The supersede is part of the document's unit of work, not its own commit.

    An Annual Leave renders a ``Leave Undertaking`` companion in step 14, i.e.
    after the hook. Failing there must take the leave row *and* the absence
    cleanup back with it; with the hook committing on its own, the absence is
    already gone and the operator is left with neither paper nor mark.
    """

    real_corner = generation_env.aztec_corner_for

    def boom(template_id: str) -> str:
        if template_id == "Leave Undertaking":
            raise RuntimeError("post-step-12 failure")
        return real_corner(template_id)

    monkeypatch.setattr(generation_env, "aztec_corner_for", boom)

    _guard(db_session)
    client.put(
        "/api/v1/timesheet/2026/7/cell", json={"employee_id": "G1001", "day": 9, "code": "AB"}
    )
    assert db_session.query(Absence).count() == 1

    with pytest.raises(RuntimeError, match="post-step-12"):
        generation_env.generate_document(
            db_session,
            employee_id="G1001",
            template_id="Leave Application Form",
            fields={
                "leave_type": "Annual Leave",
                "start_date": "2026-07-09",
                "end_date": "2026-07-09",
                "total_days": 1,
            },
            commit=True,
        )
    db_session.rollback()
    assert db_session.query(Absence).count() == 1
    assert db_session.query(Leave).count() == 0
    assert client.get("/api/v1/timesheet/2026/7").json()["rows"][0]["codes"][8] == "AB"
