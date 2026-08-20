"""The rendered workbook must match the paper already in circulation.

Every assertion here reopens the produced bytes with openpyxl and reads the real
workbook — values, formulas, merges, row heights, the conditional-format XML and
the data validation — because the deliverable is paper the client already holds
and a byte count proves nothing about it.

No database: ``MonthGrid`` and ``GridRow`` are frozen dataclasses, so the grids
are built directly. The renderer takes no ``Session`` and this module must never
need one.
"""

from __future__ import annotations

import hashlib
import io
import shutil
from collections.abc import Iterator, Mapping
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from app.api.errors import NotFoundError
from app.config import get_settings
from app.core import timesheet_xlsx
from app.core.timesheet_codes import CODE_PRESENT, CODE_SICK
from app.services.timesheet_service import CODE_BLOCKED, GridRow, MonthGrid

# --------------------------------------------------------------------------- #
# Grids, built without a database
# --------------------------------------------------------------------------- #


def _codes(
    days: int, *, base: str = CODE_PRESENT, on: Mapping[int, str] | None = None
) -> list[str | None]:
    """One month of day codes: ``base`` up to ``days``, ``None`` past the end."""

    cells: list[str | None] = [base if day <= days else None for day in range(1, 32)]
    for day, code in (on or {}).items():
        cells[day - 1] = code
    return cells


def _row(
    employee_id: str = "G1001",
    row_no: int = 1,
    *,
    name_en: str = "TEST GUARD",
    days: int = 31,
    codes: list[str | None] | None = None,
    stat_codes: list[str | None] | None = None,
    stat_block: int = 1,
) -> GridRow:
    day_codes = _codes(days) if codes is None else codes
    return GridRow(
        employee_id=employee_id,
        row_no=row_no,
        name_en=name_en,
        nationality_en="U.A.E",
        designation_en="Security Guard",
        designation_ar="حارس امن",
        rank_order=15,
        codes=day_codes,
        stat_codes=list(day_codes) if stat_codes is None else stat_codes,
        stat_block=stat_block,
        stat_filler=None,
        joined_day=None,
        left_day=None,
        start_confirmed=True,
        notes={},
    )


def _grid(
    *rows: GridRow,
    year: int = 2026,
    month: int = 7,
    days: int = 31,
    sheet: str = "main",
    post_count: int = 249,
) -> MonthGrid:
    return MonthGrid(
        year=year,
        month=month,
        days_in_month=days,
        sheet=sheet,
        post_count=post_count,
        rows=list(rows),
        blocking=[],
        warnings=[],
        removed=[],
        closed_at=None,
        closed_by=None,
    )


def _sheet(payload: bytes) -> Any:
    return load_workbook(io.BytesIO(payload)).worksheets[0]


def _template() -> Path:
    return get_settings().templates_dir / "GSSG-HR_Monthly_Time_Sheet.xlsx"


# --------------------------------------------------------------------------- #
# The attendance sheet
# --------------------------------------------------------------------------- #


def test_attendance_sheet_keeps_the_logo_and_header() -> None:
    sheet = _sheet(timesheet_xlsx.render(_grid(_row())))
    assert len(sheet._images) == 1
    assert sheet["D4"].value == "For the Month of :JUL-2026"
    assert sheet["Q2"].value == " Site Name :   JD 908"
    assert sheet.freeze_panes == "F6"


def test_a_data_row_carries_values_and_the_countif_formulas() -> None:
    sheet = _sheet(timesheet_xlsx.render(_grid(_row())))
    assert [sheet.cell(6, c).value for c in (1, 2, 3, 4, 5)] == [
        1,
        "G1001",
        "TEST GUARD",
        "U.A.E",
        "Security Guard",
    ]
    assert sheet["F6"].value == "P"
    assert sheet["AJ6"].value == "P"
    assert sheet["AK6"].value == '=COUNTIF(F6:AJ6,"P")'
    assert sheet["AL6"].value == '=COUNTIF(F6:AJ6,"OFF")'
    assert sheet["AM6"].value == '=COUNTIF(F6:AJ6,"AB")'
    assert sheet["AN6"].value == '=COUNTIF(F6:AK6,"AL")'  # spans AK in the original
    assert sheet["AO6"].value == "=COUNTIF(F6:AJ6,$AO$5)"  # $AO$5 holds "SL "
    assert sheet["AP6"].value == '=COUNTIF(F6:AJ6,"TR")'


def test_a_thirty_day_month_leaves_day_31_empty() -> None:
    sheet = _sheet(timesheet_xlsx.render(_grid(_row(days=30), month=6, days=30)))
    assert sheet["AI6"].value == "P"
    assert sheet["AJ6"].value is None
    assert sheet["D4"].value == "For the Month of :JUN-2026"


def test_row_heights_come_from_the_template() -> None:
    specimen = load_workbook(_template())["_parts"].row_dimensions[1].height
    sheet = _sheet(timesheet_xlsx.render(_grid(_row(), _row("G1002", 2))))
    assert sheet.row_dimensions[6].height == specimen
    assert sheet.row_dimensions[7].height == specimen
    # The footer's own heights ride along with the copied block.
    assert sheet.row_dimensions[14].height == 35.1  # the S.no header row
    assert sheet.row_dimensions[25].height == 35.1  # X / Not billed, a code row
    assert sheet.row_dimensions[26].height == 39.95  # Total Days


# --------------------------------------------------------------------------- #
# The footer
# --------------------------------------------------------------------------- #


def test_the_footer_follows_the_last_data_row() -> None:
    """One data row: L=6. Legend 7, signatures 8, sums 9, three blanks 10-12,
    S.no 13, the ELEVEN code rows 14-24, Total Days 25. The eleventh code row is
    the red block X, which is why Total Days is 25 and not June's 24."""

    sheet = _sheet(timesheet_xlsx.render(_grid(_row())))
    assert "Legend:" in str(sheet["A7"].value)
    assert "X- Not billed" in str(sheet["A7"].value)
    assert str(sheet["A8"].value).startswith("Prepard By")
    assert sheet["N8"].value == "Verfied By "
    assert sheet["AD8"].value == "Approved By "
    assert sheet["AK9"].value == "=SUM(AK6:AK6)"
    assert sheet["AP9"].value == "=SUM(AP6:AP6)"
    assert [sheet.cell(r, 1).value for r in (10, 11, 12)] == [None, None, None]
    assert sheet["A13"].value == "S.no"
    assert sheet["C13"].value == "STATE"
    assert sheet["D13"].value == "CODE"
    assert sheet["A14"].value == "Total "
    assert sheet["D24"].value == "X"  # the red block, above Total Days
    assert sheet["A25"].value == "Total Days"
    assert sheet["E25"].value == "=SUM(E14:E24)"


def test_the_footer_code_rows_carry_the_normalised_formulas() -> None:
    """The four deliberate divergences from June live here.

    June has OFF as ``=COUNTIF(F7:AJ288,"OFF")`` — off by one at both ends — and
    ``-``/R/S as ``=COUNTIF(F6:AJ287,D300)`` style cell references, with R and S
    stopping 24 rows short of the roster. All four are normalised.
    """

    sheet = _sheet(timesheet_xlsx.render(_grid(_row(), _row("G1002", 2))))
    codes = [sheet.cell(r, 4).value for r in range(15, 26)]
    assert codes == ["SL ", "AL", "AB", "TR", "NG", "-", "R", "S ", "P", "OFF", "X"]
    formulas = [sheet.cell(r, 5).value for r in range(15, 26)]
    assert formulas == [
        "=AO10",  # Sick Leave
        "=AN10",  # Annual Leave
        "=AM10",  # Abcent
        "=AP10",  # National Service
        '=COUNTIF(F6:AJ7,"NG")',  # New Gard
        '=COUNTIF(F6:AJ7,"-")',  # Termination — a literal, not D300
        '=COUNTIF(F6:AJ7,"R")',  # Resignation — full roster, not 24 rows short
        '=COUNTIF(F6:AJ7,"S ")',  # Suspention — the code keeps its trailing space
        "=AK10",  # P
        "=AL10",  # OFF — not June's off-by-one COUNTIF
        '=COUNTIF(F6:AJ7,"X")',  # Not billed
    ]
    assert sheet["E26"].value == "=SUM(E15:E25)"


def test_the_footer_merges_are_reapplied() -> None:
    sheet = _sheet(timesheet_xlsx.render(_grid(_row())))
    merges = {str(m) for m in sheet.merged_cells.ranges}
    assert {"A7:AP7", "A8:M8", "N8:AC8", "AD8:AP8", "A13:B13", "A14:B24", "A25:D25"} <= merges
    # rows 1-5 keep the template's own merges
    assert "D4:AH4" in merges


def test_an_empty_roster_keeps_the_ranges_well_formed() -> None:
    """A drivers month with nobody on it still has to be a valid workbook.

    The band never collapses to nothing, because ``=SUM(AK6:AK5)`` is a reversed
    reference that would fold the header row into the total.
    """

    sheet = _sheet(timesheet_xlsx.render(_grid()))
    assert sheet["B6"].value is None
    assert sheet["AK9"].value == "=SUM(AK6:AK6)"
    assert sheet["A7"].value is not None and "Legend:" in str(sheet["A7"].value)


# --------------------------------------------------------------------------- #
# Conditional formatting and the code validation
# --------------------------------------------------------------------------- #


def test_conditional_formats_span_only_the_real_extent() -> None:
    sheet = _sheet(timesheet_xlsx.render(_grid(_row())))
    ranges = {str(cf.sqref) for cf in sheet.conditional_formatting}
    assert ranges == {"F6:AJ6"}


def test_conditional_format_fills_use_bgcolor_and_the_spaced_sick_code() -> None:
    """The source stores its dxf fills on ``bgColor`` with no ``patternType``.

    ``PatternFill(start_color=..., fill_type="solid")`` writes ``fgColor``
    instead and produces different XML. The sick rule must test ``"SL "`` WITH
    the trailing space — the source tests ``"SL"`` without it, which is why the
    rule never fires on the hand file.
    """

    sheet = _sheet(timesheet_xlsx.render(_grid(_row())))
    rules = next(iter(sheet.conditional_formatting)).rules
    produced = {}
    for rule in rules:
        assert rule.type == "cellIs"
        assert rule.operator == "equal"
        assert rule.dxf.fill.patternType is None
        assert rule.dxf.fill.fgColor.rgb == "00000000"  # untouched default
        font = rule.dxf.font
        produced[rule.formula[0]] = (
            rule.dxf.fill.bgColor.rgb,
            None if font is None else font.color.rgb,
        )
    assert produced == {
        '"AL"': ("FFBDD7EE", None),
        '"SL "': ("FFC6E0B4", None),
        '"AB"': ("FFFFC7CE", "FF9C0006"),
        '"TR"': ("FFCC99FF", None),
        '"NG"': ("FFFF9900", None),
        '"X"': ("FF990033", "FFFFFFFF"),
    }
    assert f'"{CODE_SICK}"' in produced  # the constant is the spaced one
    assert f'"{CODE_BLOCKED}"' in produced


def test_the_code_validation_is_a_quoted_literal_list() -> None:
    """The source points its three list validations at ``$D$295:$D$304``, a
    reference into the footer code rows. The renderer replaces that with a
    literal list, which openpyxl needs quoted."""

    sheet = _sheet(timesheet_xlsx.render(_grid(_row(), _row("G1002", 2))))
    validations = sheet.data_validations.dataValidation
    assert len(validations) == 1
    validation = validations[0]
    assert validation.type == "list"
    assert validation.formula1 == '"P,AL,SL ,AB,TR,NG,-,X"'
    assert validation.allow_blank is True
    assert str(validation.sqref) == "F6:AJ7"


# --------------------------------------------------------------------------- #
# The statistics variant
# --------------------------------------------------------------------------- #


def test_statistics_uses_arabic_designations() -> None:
    sheet = _sheet(timesheet_xlsx.render(_grid(_row()), variant="statistics"))
    assert sheet["E6"].value == "حارس امن"
    # ...where the attendance variant prints the English one.
    assert _sheet(timesheet_xlsx.render(_grid(_row())))["E6"].value == "Security Guard"


def test_statistics_prints_stat_codes_not_the_real_ones() -> None:
    row = _row(codes=_codes(31, on={4: "AB"}), stat_codes=_codes(31))
    sheet = _sheet(timesheet_xlsx.render(_grid(row), variant="statistics"))
    assert sheet["I6"].value == "P"
    assert _sheet(timesheet_xlsx.render(_grid(row)))["I6"].value == "AB"


def test_statistics_splits_blocks_with_two_blank_rows() -> None:
    """``post_count = 0`` — everyone is surplus, so block 1 is empty and the gap
    still comes first."""

    grid = _grid(_row(stat_block=2), post_count=0)
    sheet = _sheet(timesheet_xlsx.render(grid, variant="statistics"))
    assert sheet["B6"].value is None and sheet["B7"].value is None
    assert sheet["B8"].value == "G1001"
    assert sheet["A8"].value == 1  # numbering continues across the gap
    assert sheet["AK8"].value == '=COUNTIF(F8:AJ8,"P")'
    assert sheet["AK7"].value is None  # the gap rows carry no formulas
    assert sheet["AK11"].value == "=SUM(AK6:AK8)"  # the sums span the gap


def test_statistics_keeps_the_gap_between_the_two_blocks() -> None:
    grid = _grid(
        _row("G1001", 1),
        _row("G1002", 2),
        _row("G1003", 3, stat_block=2),
        post_count=2,
    )
    sheet = _sheet(timesheet_xlsx.render(grid, variant="statistics"))
    assert [sheet.cell(r, 2).value for r in range(6, 11)] == [
        "G1001",
        "G1002",
        None,
        None,
        "G1003",
    ]
    assert sheet["A10"].value == 3  # the roster position, not the sheet row
    assert sheet["A11"].value is not None and "Legend:" in str(sheet["A11"].value)


def test_attendance_never_splits_the_roster() -> None:
    grid = _grid(_row("G1001", 1), _row("G1002", 2, stat_block=2), post_count=1)
    sheet = _sheet(timesheet_xlsx.render(grid))
    assert [sheet.cell(r, 2).value for r in (6, 7)] == ["G1001", "G1002"]
    assert "Legend:" in str(sheet["A8"].value)


def test_an_unknown_variant_is_rejected() -> None:
    with pytest.raises(ValueError, match="variant"):
        timesheet_xlsx.render(_grid(_row()), variant="statistcs")
    with pytest.raises(ValueError, match="variant"):
        timesheet_xlsx.filename_for(_grid(_row()), variant="nonsense")


# --------------------------------------------------------------------------- #
# File names
# --------------------------------------------------------------------------- #


def test_filenames_are_the_arabic_names_in_use() -> None:
    grid = _grid(_row())
    assert timesheet_xlsx.filename_for(grid) == "كشف حضور شهر يوليو.xlsx"
    assert timesheet_xlsx.filename_for(grid, variant="statistics") == "الاحصائية شهر يوليو.xlsx"
    assert timesheet_xlsx.filename_for_single(grid, "G1001") == "كشف حضور TEST GUARD يوليو.xlsx"


def test_drivers_filename_has_its_own_suffix() -> None:
    grid = _grid(_row(), sheet="drivers")
    assert timesheet_xlsx.filename_for(grid) == "كشف حضور شهر يوليو للسائقين.xlsx"
    assert (
        timesheet_xlsx.filename_for(grid, variant="statistics")
        == "الاحصائية شهر يوليو للسائقين.xlsx"
    )


# --------------------------------------------------------------------------- #
# The single-employee extracts
# --------------------------------------------------------------------------- #


def test_a_single_employee_sheet_has_one_row() -> None:
    """Three on the roster, one on the paper — the middle one."""

    grid = _grid(_row("G1001", 1), _row("G1002", 2), _row("G1003", 3))
    sheet = _sheet(timesheet_xlsx.render_single(grid, "G1002"))
    assert sheet["B6"].value == "G1002"
    assert sheet["A6"].value == 2  # his position on the roster, kept
    assert sheet["B7"].value is None
    assert "Legend:" in str(sheet["A7"].value)
    assert sheet["AK9"].value == "=SUM(AK6:AK6)"


def test_a_two_month_extract_carries_a_sheet_per_month() -> None:
    """The resignation and termination handover: month of departure + the one before."""

    grids = [
        _grid(_row("G1001", 1, days=30), _row("G1002", 2, days=30), month=6, days=30),
        _grid(_row("G1001", 1), _row("G1002", 2), month=7),
    ]
    workbook = load_workbook(io.BytesIO(timesheet_xlsx.render_single_span(grids, "G1001")))
    assert workbook.sheetnames == ["JUN", "JUL"]
    assert workbook["JUN"]["B6"].value == "G1001"
    assert workbook["JUL"]["D4"].value == "For the Month of :JUL-2026"
    assert workbook["JUN"]["D4"].value == "For the Month of :JUN-2026"
    for name in ("JUN", "JUL"):
        sheet = workbook[name]
        assert len(sheet._images) == 1, name
        assert sheet.freeze_panes == "F6"
        assert sheet["B7"].value is None
        assert "Legend:" in str(sheet["A7"].value)
        assert sheet["AK9"].value == "=SUM(AK6:AK6)"
        assert str(sheet.data_validations.dataValidation[0].sqref) == "F6:AJ6"
    assert workbook["JUN"]["AJ6"].value is None  # June has 30 days
    assert workbook["JUL"]["AJ6"].value == "P"


def test_a_span_month_the_employee_missed_is_an_empty_sheet() -> None:
    """Joined in July: June is his sheet with nothing on it, not a failure."""

    grids = [_grid(month=6, days=30), _grid(_row("G1001", 1), month=7)]
    workbook = load_workbook(io.BytesIO(timesheet_xlsx.render_single_span(grids, "G1001")))
    assert workbook.sheetnames == ["JUN", "JUL"]
    assert workbook["JUN"]["B6"].value is None
    assert workbook["JUL"]["B6"].value == "G1001"


def test_an_employee_off_the_sheet_is_not_found() -> None:
    grid = _grid(_row("G1001", 1))
    with pytest.raises(NotFoundError):
        timesheet_xlsx.render_single(grid, "G9999")
    with pytest.raises(NotFoundError):
        timesheet_xlsx.filename_for_single(grid, "G9999")
    with pytest.raises(NotFoundError):
        timesheet_xlsx.render_single_span([grid], "G9999")


def test_a_span_needs_at_least_one_month() -> None:
    with pytest.raises(ValueError, match="month"):
        timesheet_xlsx.render_single_span([], "G1001")


# --------------------------------------------------------------------------- #
# The template itself
# --------------------------------------------------------------------------- #


def test_the_helper_sheet_is_deleted() -> None:
    workbook = load_workbook(io.BytesIO(timesheet_xlsx.render(_grid(_row()))))
    assert workbook.sheetnames == ["Sheet1"]


def test_the_template_file_is_never_rewritten() -> None:
    before = hashlib.sha256(_template().read_bytes()).hexdigest()
    timesheet_xlsx.render(_grid(_row()))
    timesheet_xlsx.render(_grid(_row()), variant="statistics")
    timesheet_xlsx.render_single_span([_grid(_row()), _grid(_row(), month=8)], "G1001")
    assert hashlib.sha256(_template().read_bytes()).hexdigest() == before


@pytest.fixture()
def _clean_settings() -> Iterator[None]:
    get_settings.cache_clear()
    yield
    get_settings.cache_clear()


def test_the_template_is_resolved_at_call_time(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, _clean_settings: None
) -> None:
    """``GSSG_TEMPLATES_DIR`` moves the template, so the path cannot be a
    module-level ``__file__`` constant — the PyInstaller layout puts it under
    ``sys._MEIPASS/templates/``."""

    source = _template()
    empty = tmp_path / "empty"
    empty.mkdir()
    monkeypatch.setenv("GSSG_TEMPLATES_DIR", str(empty))
    get_settings.cache_clear()
    with pytest.raises(FileNotFoundError):
        timesheet_xlsx.render(_grid(_row()))

    moved = tmp_path / "moved"
    moved.mkdir()
    shutil.copy2(source, moved / source.name)
    monkeypatch.setenv("GSSG_TEMPLATES_DIR", str(moved))
    get_settings.cache_clear()
    assert _sheet(timesheet_xlsx.render(_grid(_row())))["B6"].value == "G1001"
