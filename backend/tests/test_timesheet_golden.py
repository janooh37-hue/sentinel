"""Reproduce the day codes on the hand-kept June and July 2026 sheets.

Skipped unless the finance share and the live DB are both reachable — this is a
site-specific acceptance gate, not a portable unit test. It is the only thing in
the repo that regression-guards these numbers: the import script's ``--verify``
mode merely *prints* a cell diff and asserts nothing, so this module is what
turns that print into a guard.

**What is compared.** The day-code band of each hand workbook — every employee
row, every day column — keyed by employee ID, against
:func:`app.services.timesheet_service.build_month`. That covers the roster rules
(who is on which sheet for the month), the day-code rules in
:mod:`app.core.timesheet_codes` resolved against the live leaves, absences and
roster edges, the two statistics blocks' row split, and the preflight's
``blocking`` verdict. Those are the parts a wrong answer would silently misprice
on the client's invoice.

**What is not compared.** Row *order*: the comparison is keyed by ID and is
order-insensitive by design, because the engine sorts by ``rank_order`` then
numeric ID and deliberately does not reproduce the paper's guard-tier order
(``docs/superpowers/specs/2026-08-19-monthly-timesheet-design.md:151, :295, :507-509``) — only 36 of July's 275 row positions
coincide. Nor the names, nationalities or designation labels beside the band;
nor the rendered ``.xlsx``, the template or the API, none of which this module
imports — its only production imports are :func:`attach_sqlite_pragmas`,
``CODE_ABSENT`` and ``timesheet_service``. "0 differing cells of 8,525" means
zero wrong day codes per man, not a page that overlays the paper.

The baseline, counted off the paper the client already holds:

* **July: 0 differing cells of 8,525** — 275 rows x 31 days, exact.
* **June: 49 differing cells of 8,460** — 282 rows x 30 days, and every one of
  the 49 is a departure-date error in the *hand* file rather than an engine
  error. Six employees, all of whom the paper keeps on the wrong number of days:

  - ``G3808``  end_date 2026-04-02, **30** cells: the paper has him present all
    month, but he finished in April, so the engine does not put him on the June
    roster at all.
  - ``G4053``  end_date 2026-06-17, **13** cells: the paper runs him to the 30th.
  - ``G0984``  end_date 2026-06-08, **3** cells: the paper drops him after the 5th.
  - ``G3636``  end_date 2026-06-08, **1** cell: the paper drops him after the 7th.
  - ``G3699``  end_date 2026-06-08, **1** cell: the paper drops him after the 7th.
  - ``G4532``  end_date 2026-06-06, **1** cell: the paper drops him after the 5th.

A row on the paper that the engine never generates is compared against blanks,
not skipped, and so costs a full row of differences — that is where ``G3808``'s
30 come from. Skipping it instead would leave 30 cells of silent headroom in
June's budget, which is the whole failure this module exists to prevent: a diff
of zero is only evidence when the comparison is known to have looked.

Read-only throughout. This is the production database, ``build_month`` never
writes (it creates no ``TimesheetPeriod`` row for a month that has none), and
:func:`test_the_live_session_refuses_to_write` proves the connection could not
write even if that changed.
"""

from __future__ import annotations

import calendar
from collections import Counter
from collections.abc import Iterator, Mapping
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import Session, sessionmaker

from app.core.timesheet_codes import CODE_ABSENT
from app.db.session import attach_sqlite_pragmas
from app.services import timesheet_service as svc

SHARE = Path(r"E:\Al Watbha Shares\المالية\احصائية 2026")
LIVE_DB = Path(__file__).resolve().parents[2] / "data" / "gssg.db"
SHEETS = {
    6: SHARE / r"6-Jun\كشف حضور شهر يونيو.xlsx",
    7: SHARE / r"7-Jul\كشف حضور شهر يوليو.xlsx",
}
#: Drivers have always been reported on their own workbook.
DRIVER_SHEET = SHARE / r"7-Jul\كشف حضور شهر يوليو للسائقين.xlsx"

#: The paper's geometry. Row 5 is the header, row 6 the first man; within a row
#: tuple index 1 is the ID and index 5 is day 1 (``# | ID | Name | Nat | Desig``).
HEADER_ROW = 5
ID_INDEX = 1
DAY_INDEX = 5

#: Data rows on the paper, and therefore cells compared: 282 x 30 = 8,460 for
#: June, 275 x 31 = 8,525 for July.
HAND_ROWS = {6: 282, 7: 275}

#: The differing cells, per employee. See the module docstring: June's six are
#: all departure-date errors in the hand file.
DIFFS_BY_EMPLOYEE: dict[int, dict[str, int]] = {
    6: {"G3808": 30, "G4053": 13, "G0984": 3, "G3636": 1, "G3699": 1, "G4532": 1},
    7: {},
}
ALLOWED_DIFFS = {6: 49, 7: 0}

#: Roster identity against the paper, both directions. ``G3808`` finished in
#: April and belongs on neither sheet; ``G4810`` and ``G5704`` were employed all
#: June and the paper simply left them out — the omission the import documented.
HAND_ONLY: dict[int, set[str]] = {6: {"G3808"}, 7: set()}
ENGINE_ONLY: dict[int, set[str]] = {6: {"G4810", "G5704"}, 7: set()}

#: ``(block 1, block 2)`` row counts. June's block 2 is 34 where the paper
#: carried 33, which is that same roster difference and nothing else:
#: 282 - 1 (G3808) + 2 (G4810, G5704) = 283 = 249 + 34.
STAT_BLOCKS = {6: (249, 34), 7: (249, 26)}

DRIVER_IDS = ("G5566", "G5567")

#: The drivers workbook's one divergence, and it is the *paper* that is wrong —
#: a one-cell slip of the same class as June's 49, not a rival convention. G5567
#: joined 2026-07-01 and the drivers sheet writes ``NG`` on his joining day;
#: ``month_codes`` writes ``NG`` strictly *before* doj and ``P`` on the day the
#: man actually starts. The paper's own main sheet says the same thing five times
#: over — every June joiner on it reads ``NG`` up to the day before and ``P`` on
#: the day: ``G5524`` (doj 06-02) one ``NG`` then ``P``; ``G5530`` and ``G5558``
#: (doj 06-08) seven then ``P``; ``G5260`` and ``G5677`` (doj 06-09) eight then
#: ``P``. Those five joining-day ``P``\ s are inside the June band this module
#: compares, so "fixing" ``month_codes`` to stamp ``NG`` on a joining day would
#: put five new differences into June and break ``DIFFS_BY_EMPLOYEE[6]``.
#: Pinned, not tolerated: if the rule ever moves, this list says so.
DRIVER_DIFFS = [("G5567", 1, "P", "NG")]

#: Every path the module opens, not just two of them: on a partially synced share
#: this must skip, never error.
pytestmark = pytest.mark.skipif(
    not LIVE_DB.exists() or not all(path.exists() for path in (*SHEETS.values(), DRIVER_SHEET)),
    reason="needs the live DB and the finance share",
)


@pytest.fixture(scope="module")
def live_session() -> Iterator[Session]:
    # Read-only: this is the production database. ``wal=False`` matters — it
    # keeps ``attach_sqlite_pragmas`` to ``foreign_keys=ON`` and off the
    # ``journal_mode`` write a read-only file would refuse anyway.
    engine = create_engine(
        f"sqlite:///file:{LIVE_DB.as_posix()}?mode=ro&uri=true",
        future=True,
        connect_args={"uri": True},
    )
    attach_sqlite_pragmas(engine, wal=False)
    session = sessionmaker(bind=engine, future=True, expire_on_commit=False)()
    try:
        yield session
    finally:
        session.close()
        engine.dispose()


def _hand_sheet(path: Path, days: int) -> dict[str, list[str]]:
    """The paper's data band: employee ID -> ``days`` upper-cased day codes.

    Stops at the first blank ID, where the band ends and the legend, the
    statistics footer and the signature block begin. Counting one of those as a
    data row would manufacture differences that are not on the paper.

    The header row is checked rather than trusted, because every cell address
    below depends on it and a shifted column would read as a wall of real diffs.
    """

    workbook = load_workbook(path, read_only=True)
    try:
        rows = workbook.worksheets[0].iter_rows(
            min_row=HEADER_ROW, max_col=DAY_INDEX + days, values_only=True
        )
        header = next(rows)
        assert header[ID_INDEX] == "ID", f"{path.name}: column B is not the ID column"
        assert [header[DAY_INDEX + i] for i in range(days)] == list(range(1, days + 1)), (
            f"{path.name}: the day columns are not 1..{days}"
        )

        band: dict[str, list[str]] = {}
        for row in rows:
            employee_id = "" if row[ID_INDEX] is None else str(row[ID_INDEX]).strip()
            if not employee_id:
                break
            band[employee_id] = [
                "" if row[DAY_INDEX + i] is None else str(row[DAY_INDEX + i]).strip().upper()
                for i in range(days)
            ]
        return band
    finally:
        workbook.close()


def _compare(
    hand: Mapping[str, list[str]],
    generated: Mapping[str, list[str | None]],
    days: int,
) -> tuple[int, list[tuple[str, int, str, str]]]:
    """Every cell of the paper's band against the engine.

    Returns the number of cells examined and ``(employee, day, ours, theirs)``
    for each disagreement. A man on the paper whom the engine never generated is
    compared against blanks, so he costs a full row: "not on this roster" against
    "worked every day" is the strongest disagreement there is, not one to skip.
    """

    compared = 0
    diffs: list[tuple[str, int, str, str]] = []
    for employee_id, theirs in hand.items():
        mine = generated.get(employee_id)
        for index in range(days):
            compared += 1
            ours = "" if mine is None else (mine[index] or "").strip().upper()
            if ours != theirs[index]:
                diffs.append((employee_id, index + 1, ours, theirs[index]))
    return compared, diffs


def test_the_live_session_refuses_to_write(live_session: Session) -> None:
    """The safety the rest of this module rests on, asserted instead of promised.

    ``build_month`` runs here against the real production file. A session that
    could write is one refactor away from stamping a ``TimesheetPeriod`` row into
    it, so the connection is checked to be incapable of it.

    The one deliberate write attempt in this module runs inside a savepoint, so
    that if the read-only guarantee ever *did* break, the ``CREATE TABLE`` would
    succeed, ``pytest.raises`` would fail the test, and the savepoint would still
    undo it on the way out — containment stated by the test rather than inherited
    from fixture teardown.
    """

    with live_session.begin_nested() as savepoint:
        with pytest.raises(OperationalError, match="readonly database"):
            live_session.execute(text("CREATE TABLE _golden_probe (x INTEGER)"))
        savepoint.rollback()
    assert live_session.execute(text("PRAGMA foreign_keys")).scalar() == 1


@pytest.mark.parametrize("month", [6, 7])
def test_generated_grid_matches_the_hand_kept_sheet(live_session: Session, month: int) -> None:
    days = calendar.monthrange(2026, month)[1]
    hand = _hand_sheet(SHEETS[month], days)
    assert len(hand) == HAND_ROWS[month], f"2026-{month:02d}: read {len(hand)} rows off the paper"

    grid = svc.build_month(live_session, 2026, month)
    generated = {row.employee_id: row.codes for row in grid.rows}
    assert set(hand) - set(generated) == HAND_ONLY[month]
    assert set(generated) - set(hand) == ENGINE_ONLY[month]

    compared, diffs = _compare(hand, generated, days)
    # The count that makes a diff of zero mean something: 8,460 and 8,525.
    assert compared == HAND_ROWS[month] * days, f"2026-{month:02d}: compared {compared} cells"
    assert dict(Counter(employee for employee, *_ in diffs)) == DIFFS_BY_EMPLOYEE[month], (
        f"2026-{month:02d}: {diffs[:20]}"
    )
    assert len(diffs) == ALLOWED_DIFFS[month], f"2026-{month:02d}: {len(diffs)} differing cells"


def test_the_comparison_would_notice_a_wrong_cell(live_session: Session) -> None:
    """July's zero means "looked and agreed", not "looked at nothing".

    One cell of the generated grid is perturbed in memory — nothing is written
    anywhere — and the comparison must find exactly that cell and no other. A
    comparator that quietly examined nothing would report the same zero above,
    so the zero is only evidence alongside this.
    """

    hand = _hand_sheet(SHEETS[7], 31)
    grid = svc.build_month(live_session, 2026, 7)
    generated = {row.employee_id: list(row.codes) for row in grid.rows}
    assert _compare(hand, generated, 31) == (HAND_ROWS[7] * 31, [])

    victim = grid.rows[0].employee_id
    assert hand[victim][14] != CODE_ABSENT
    generated[victim][14] = CODE_ABSENT
    assert _compare(hand, generated, 31) == (
        HAND_ROWS[7] * 31,
        [(victim, 15, CODE_ABSENT, hand[victim][14])],
    )

    # And a whole row going missing costs a whole row, which is what buys G3808
    # his 30 cells of June.
    del generated[victim]
    compared, diffs = _compare(hand, generated, 31)
    assert (compared, len(diffs)) == (HAND_ROWS[7] * 31, 31)


def test_july_roster_is_the_275_on_the_sheet(live_session: Session) -> None:
    assert len(svc.build_month(live_session, 2026, 7).rows) == HAND_ROWS[7]
    drivers = svc.build_month(live_session, 2026, 7, sheet="drivers")
    assert tuple(row.employee_id for row in drivers.rows) == DRIVER_IDS


def test_the_drivers_workbook_reproduces_its_own_sheet(live_session: Session) -> None:
    """The second deliverable, whose two men are on no main sheet."""

    hand = _hand_sheet(DRIVER_SHEET, 31)
    assert tuple(hand) == DRIVER_IDS

    drivers = svc.build_month(live_session, 2026, 7, sheet="drivers")
    generated = {row.employee_id: row.codes for row in drivers.rows}
    compared, diffs = _compare(hand, generated, 31)
    assert compared == len(DRIVER_IDS) * 31
    assert diffs == DRIVER_DIFFS


@pytest.mark.parametrize("month", [6, 7])
def test_statistics_blocks_match_the_hand_kept_split(live_session: Session, month: int) -> None:
    grid = svc.build_month(live_session, 2026, month)
    counts = Counter(row.stat_block for row in grid.rows)
    assert (counts[1], counts[2]) == STAT_BLOCKS[month]
    assert grid.post_count == STAT_BLOCKS[month][0]


def test_the_live_data_has_no_blocking_issues(live_session: Session) -> None:
    """G5678 joined 2026-08-03 with no designation — he must not block July.

    Both workbooks: a blocking issue on either one stops that download.
    """

    for sheet in svc.SHEETS:
        assert svc.build_month(live_session, 2026, 7, sheet=sheet).blocking == []
